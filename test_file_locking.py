"""
Test script to verify Excel file locking and recovery functionality on Windows.
This script intentionally creates file locking scenarios and tests the recovery mechanisms.
"""

import os
import sys
import time
import tempfile
import shutil
import logging
import threading
import argparse
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('file_lock_test.log')
    ]
)

logger = logging.getLogger("FileLockTest")

def create_test_excel_file(directory: str = None) -> str:
    """
    Create a test Excel file for locking tests.
    
    Args:
        directory: Optional directory to create the file in
        
    Returns:
        str: Path to the created Excel file
    """
    try:
        import openpyxl
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add some data
        ws['A1'] = "Test Data"
        ws['B1'] = "Value"
        for i in range(1, 11):
            ws[f'A{i+1}'] = f"Row {i}"
            ws[f'B{i+1}'] = i * 10
        
        # Create directory if specified
        if directory:
            os.makedirs(directory, exist_ok=True)
            
        # Save to a temp file
        if directory:
            temp_file = os.path.join(directory, f"test_excel_{int(time.time())}.xlsx")
        else:
            temp_file = os.path.join(tempfile.gettempdir(), f"test_excel_{int(time.time())}.xlsx")
            
        wb.save(temp_file)
        wb.close()
        
        logger.info(f"Created test Excel file at: {temp_file}")
        return temp_file
        
    except Exception as e:
        logger.error(f"Failed to create test Excel file: {e}")
        raise

def lock_file_with_excel(file_path: str, lock_duration: int = 10):
    """
    Lock an Excel file by opening it with Excel and holding it open.
    
    Args:
        file_path: Path to Excel file
        lock_duration: Duration to hold the lock in seconds
    """
    try:
        # Import required modules
        import win32com.client
        import pythoncom
        
        # Initialize COM in this thread
        pythoncom.CoInitialize()
        
        logger.info(f"Opening and locking file: {file_path}")
        
        # Start Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open the workbook
        wb = excel.Workbooks.Open(file_path)
        
        logger.info(f"File successfully locked for {lock_duration} seconds")
        
        # Hold the lock for the specified duration
        time.sleep(lock_duration)
        
        # Clean up
        wb.Close(SaveChanges=False)
        excel.Quit()
        
        # Force cleanup
        del wb
        del excel
        
        # Uninitialize COM
        pythoncom.CoUninitialize()
        
        logger.info("File lock released")
        
    except Exception as e:
        logger.error(f"Error in lock_file_with_excel: {e}")

def test_file_recovery(file_path: str):
    """
    Test the recovery mechanism on a locked file.
    
    Args:
        file_path: Path to the test Excel file
    """
    try:
        # Wait a moment for the lock to be established
        time.sleep(2)
        
        logger.info("Starting file recovery test")
        
        # Import the recovery modules
        from excel_file_recovery import fix_excel_file_in_use_error, process_with_recovery
        from win_path_handler import normalize_windows_path
        
        # Normalize the path
        file_path = normalize_windows_path(file_path)
        
        # Create output path
        recovery_dir = os.path.join(os.path.dirname(file_path), "recovery")
        os.makedirs(recovery_dir, exist_ok=True)
        
        # Run the recovery function
        logger.info("Attempting to fix the locked file")
        success, result_path = fix_excel_file_in_use_error(file_path, recovery_dir)
        
        if success:
            logger.info(f"Recovery successful: {result_path}")
            return True
        else:
            logger.error(f"Recovery failed: {result_path}")
            return False
            
    except Exception as e:
        logger.error(f"Error in test_file_recovery: {e}")
        return False

def run_locking_test(test_dir: str = None, lock_duration: int = 5):
    """
    Run a complete file locking and recovery test.
    
    Args:
        test_dir: Directory for test files
        lock_duration: Duration to hold the lock in seconds
        
    Returns:
        bool: Whether the test was successful
    """
    # Create test directory
    if not test_dir:
        test_dir = os.path.join(tempfile.gettempdir(), f"excel_test_{int(time.time())}")
        
    os.makedirs(test_dir, exist_ok=True)
    
    try:
        # Create a test Excel file
        test_file = create_test_excel_file(test_dir)
        
        # Start a thread to lock the file
        lock_thread = threading.Thread(
            target=lock_file_with_excel,
            args=(test_file, lock_duration)
        )
        lock_thread.daemon = True
        lock_thread.start()
        
        # Now try to recover the locked file
        recovery_success = test_file_recovery(test_file)
        
        # Wait for the lock thread to complete
        lock_thread.join()
        
        return recovery_success
        
    except Exception as e:
        logger.error(f"Error in run_locking_test: {e}")
        return False
    finally:
        # Clean up test files
        if os.path.exists(test_dir):
            try:
                shutil.rmtree(test_dir)
            except:
                pass

def main():
    """Main entry point for the test script."""
    parser = argparse.ArgumentParser(description="Test Excel file locking and recovery")
    parser.add_argument("--test-dir", help="Directory for test files")
    parser.add_argument("--lock-duration", type=int, default=5, help="Duration to hold the lock in seconds")
    
    args = parser.parse_args()
    
    # Add the current directory to sys.path
    current_dir = os.path.dirname(os.path.abspath(__file__))
    if current_dir not in sys.path:
        sys.path.insert(0, current_dir)
    
    # Run the test
    logger.info("Starting file locking and recovery test")
    success = run_locking_test(args.test_dir, args.lock_duration)
    
    if success:
        logger.info("TEST PASSED: File recovery successful")
        return 0
    else:
        logger.error("TEST FAILED: File recovery unsuccessful")
        return 1

if __name__ == "__main__":
    # Run only on Windows
    if sys.platform.startswith("win"):
        sys.exit(main())
    else:
        print("This test is only applicable on Windows systems")
        sys.exit(0)