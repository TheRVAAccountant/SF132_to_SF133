import openpyxl
import shutil
import logging
import openpyxl.formatting
import openpyxl.styles
import xlwings as xw
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
import sys
import traceback
import time
import os
import psutil

# Set up logging
logging.basicConfig(
    filename='file_operations.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def log_exception(e):
    """Log the exception details."""
    logging.error("Exception occurred", exc_info=True)
    print(f"An error occurred: {e}")

def close_existing_excel_instances():
    """Terminate all existing Excel processes to prevent file locking."""
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] and 'EXCEL.EXE' in proc.info['name'].upper():
            try:
                proc.terminate()
                proc.wait(timeout=5)
                logging.info(f"Terminated existing Excel process: PID {proc.pid}")
            except Exception as e:
                logging.error(f"Failed to terminate Excel process PID {proc.pid}: {e}")

def save_with_retry(wb, new_file, retries=3, delay=5):
    """Attempt to save the workbook with retries."""
    for attempt in range(retries):
        try:
            logging.info(f"Attempt {attempt + 1} to save workbook {new_file}.")
            wb.save()
            logging.info(f"Workbook {new_file} saved successfully on attempt {attempt + 1}.")
            return
        except Exception as e:
            log_exception(e)
            if attempt < retries - 1:
                logging.info(f"Retrying to save workbook after {delay} seconds.")
                time.sleep(delay)
            else:
                logging.error(f"Failed to save workbook {new_file} after {retries} attempts.")
                raise

def find_column_indexes(sheet, header_row, headers):
    """
    Find the column indexes for the specified headers by searching the header row.

    Args:
        sheet (Worksheet): The openpyxl worksheet object.
        header_row (int): The row number where headers are located.
        headers (list): List of header names to search for.

    Returns:
        dict: A dictionary mapping header names to their respective column indexes.
    """
    column_indexes = {}
    for cell in sheet[header_row]:
        if cell.value in headers:
            column_indexes[cell.value] = cell.column
    missing_headers = [header for header in headers if header not in column_indexes]
    if missing_headers:
        error_message = f"Missing headers in row {header_row}: {', '.join(missing_headers)}"
        logging.error(error_message)
        raise ValueError(error_message)
    return column_indexes

def main():
    try:
        # Close existing Excel instances to prevent file locking
        close_existing_excel_instances()

        # Define file paths
        original_file = "10-18-2024 CWMD PD12 SEP-24 SF132-SF133Recon-review.xlsx"
        new_file = "WMD FY24 Q4 SF132to133 Recon - DO.xlsx"

        # Validate file paths
        if not os.path.exists(original_file):
            logging.error(f"Original file {original_file} does not exist.")
            sys.exit(1)

        destination_dir = os.path.dirname(new_file) or '.'
        if not os.path.exists(destination_dir):
            logging.error(f"Destination directory {destination_dir} does not exist.")
            sys.exit(1)

        if len(new_file) > 260:
            logging.error(f"Destination file path {new_file} exceeds the maximum length.")
            sys.exit(1)

        # Copy and rename the workbook
        try:
            shutil.copyfile(original_file, new_file)
            logging.info(f"Copied {original_file} to {new_file}")
        except Exception as e:
            log_exception(e)
            sys.exit(1)

        # Load the workbook using openpyxl
        try:
            workbook = openpyxl.load_workbook(new_file)
            logging.info(f"Loaded workbook {new_file}")
        except Exception as e:
            log_exception(e)
            sys.exit(1)

        # Navigate to the "SF132 to SF133 Reconciliation" worksheet
        sheet_name = "SF132 to SF133 Reconciliation"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logging.info(f"Accessed sheet: {sheet_name}")
        else:
            error_message = f"Sheet {sheet_name} not found in the workbook."
            logging.error(error_message)
            raise ValueError(error_message)

        # Unprotect the sheet using a hardcoded password
        try:
            password = "LOCK"
            sheet.protection.set_password(password)
            sheet.protection.sheet = False
            logging.info(f"Unprotected sheet: {sheet_name}")
        except Exception as e:
            log_exception(e)
            raise

        # Unhide all columns in the sheet
        try:
            for col in sheet.columns:
                cell = col[0]
                if not isinstance(cell, MergedCell):
                    col_letter = get_column_letter(cell.column)
                    sheet.column_dimensions[col_letter].hidden = False
            logging.info("Unhid all columns in the sheet.")
        except Exception as e:
            log_exception(e)
            raise

        # Unmerge all merged cells
        try:
            merged_cells = list(sheet.merged_cells)
            for merged_cell in merged_cells:
                sheet.unmerge_cells(str(merged_cell))
            logging.info("Unmerged all merged cells.")
        except Exception as e:
            log_exception(e)
            raise

        # Find the header row starting on row 9
        header_row = 9

        # Define the headers to search for
        headers_to_find = ["Difference", "Include in CFO Cert Letter", "Explanation"]

        # Find column indexes for the required headers
        try:
            column_indexes = find_column_indexes(sheet, header_row, headers_to_find)
            difference_col = column_indexes["Difference"]
            include_cfo_col = column_indexes["Include in CFO Cert Letter"]
            explanation_col = column_indexes["Explanation"]
            logging.info(f"Identified columns - Difference: {difference_col}, Include in CFO Cert Letter: {include_cfo_col}, Explanation: {explanation_col}")
        except Exception as e:
            log_exception(e)
            raise

        # Obtain the fill color of the header cell
        header_cell = sheet.cell(row=header_row, column=1)
        fill_color = header_cell.fill.start_color.index

        # Convert the fill color to its RGB value
        if isinstance(fill_color, int):
            fill_color = f"{fill_color:06X}"  # Convert integer to hex string
        if fill_color in COLOR_INDEX:
            rgb_color = COLOR_INDEX[int(fill_color, 16)]
        else:
            rgb_color = fill_color  # If it's already an RGB value

        # Ensure the RGB color is in the correct format
        if not rgb_color.startswith('FF') and len(rgb_color) == 8:
            rgb_color = rgb_color[2:]

        # Log the RGB color
        color_message = f"Header cell fill color (RGB): #{rgb_color}"
        print(color_message)
        logging.info(color_message)

        # Iterate through the rows after the header row to find the first row with the same fill color
        matching_row = None
        try:
            for row in sheet.iter_rows(min_row=header_row + 1):
                cell = row[0]
                cell_fill_color = cell.fill.start_color.index
                if isinstance(cell_fill_color, int):
                    cell_fill_color = f"{cell_fill_color:06X}"  # Convert integer to hex string
                if cell_fill_color in COLOR_INDEX:
                    cell_rgb_color = COLOR_INDEX[int(cell_fill_color, 16)]
                else:
                    cell_rgb_color = cell_fill_color  # If it's already an RGB value

                # Ensure the RGB color is in the correct format
                if not cell_rgb_color.startswith('FF') and len(cell_rgb_color) == 8:
                    cell_rgb_color = cell_rgb_color[2:]

                if cell_rgb_color == rgb_color:
                    matching_row = cell.row
                    break
        except Exception as e:
            log_exception(e)
            raise

        if matching_row:
            match_message = f"First matching row: {matching_row}, Color: #{cell_rgb_color}"
            print(match_message)
            logging.info(match_message)
            dataframe_range = f"Header row: {header_row}, Dataframe range: {header_row + 1} to {matching_row - 1}"
            print(dataframe_range)
            logging.info(dataframe_range)
        else:
            no_match_message = "No matching row found."
            print(no_match_message)
            logging.info(no_match_message)

        # Add a column header "DO Comments" in the cell immediately after the last populated column on the header row
        try:
            last_col = sheet.max_column
            new_header_cell = sheet.cell(row=header_row, column=last_col + 1)
            new_header_cell.value = "DO Comments"

            # Format the new header cell
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_font = Font(color="FF0000", bold=True, size=11, name="Calibri")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Combine all alignment properties in a single Alignment object
            cell_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

            # Apply all formatting at once
            new_header_cell.fill = yellow_fill
            new_header_cell.font = red_font
            new_header_cell.border = thin_border
            new_header_cell.alignment = cell_alignment  # Single alignment setting with all properties

            new_header_column = get_column_letter(new_header_cell.column)
            sheet.column_dimensions[new_header_column].width = 25

            # Apply red_font to the entire new header column
            for cell in sheet[new_header_column]:
                cell.font = red_font

            logging.info(f"Added and formatted 'DO Comments' column at {new_header_column}.")

        except Exception as e:
            log_exception(e)
            raise

        # Print a sample of the first 5 rows of the dataframe
        try:
            for row in sheet.iter_rows(min_row=header_row, max_row=header_row + 5, max_col=last_col + 1):
                row_values = [cell.value for cell in row]
                print(row_values)
                logging.info(row_values)
            logging.info("Printed sample of the first 5 rows of the dataframe.")
        except Exception as e:
            log_exception(e)
            raise

        # Save the changes using openpyxl
        try:
            workbook.save(new_file)
            logging.info(f"Saved changes to {new_file} using openpyxl.")
        except Exception as e:
            log_exception(e)
            raise

        # Add a short delay to ensure the file system has released the file
        time.sleep(2)

        # Use xlwings to evaluate formulas and print values to the log
        app = xw.App(visible=False)
        try:
            logging.info(f"Opening workbook {new_file} with xlwings.")
            wb = app.books.open(new_file)
            ws = wb.sheets[sheet_name]
            logging.info(f"Workbook {new_file} opened successfully with xlwings.")

            for row in range(header_row + 1, matching_row):
                try:
                    difference_value = ws.range((row, difference_col)).value
                    include_cfo_value = ws.range((row, include_cfo_col)).value
                    explanation_value = ws.range((row, explanation_col)).value

                    if difference_value not in (None, ""):
                        if include_cfo_value == "N" and explanation_value not in (None, 0, ""):
                            ws.range((row, last_col + 1)).value = "Explanation Reasonable"
                        elif include_cfo_value == "Y" and explanation_value not in (None, ""):
                            ws.range((row, last_col + 1)).value = "Explanation Reasonable; Include in CFO Cert Letter"

                        log_message = (
                            f"Row {row}: Difference = {difference_value}, "
                            f"Include in CFO Letter = {include_cfo_value}, "
                            f"Explanation = {explanation_value}"
                        )
                        print(log_message)
                        logging.info(log_message)
                except Exception as row_e:
                    log_exception(row_e)
                    continue  # Continue processing the next row

        except Exception as e:
            log_exception(e)
            raise
        finally:
            try:
                if 'wb' in locals():
                    logging.info(f"Attempting to save workbook {new_file} with xlwings.")
                    save_with_retry(wb, new_file)
                    logging.info(f"Workbook {new_file} saved successfully with xlwings.")
                    wb.close()
                    logging.info(f"Workbook {new_file} closed successfully with xlwings.")
            except Exception as save_close_e:
                log_exception(save_close_e)
            try:
                app.quit()
                logging.info("Quit xlwings App successfully.")
            except Exception as app_e:
                log_exception(app_e)

    except Exception as e:
        log_exception(e)
        sys.exit(1)

if __name__ == "__main__":
    main()
