"""
Utility script to diagnose issues with Excel files.
"""
import os
import sys
import time
import logging
from pathlib import Path
import openpyxl
import win32com.client
import pythoncom
import psutil
import pywintypes

# Add parent directory to path
parent_dir = str(Path(__file__).resolve().parent.parent)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("excel_diagnosis.log")
    ]
)
logger = logging.getLogger("ExcelDiagnosis")

def analyze_excel_file(file_path: str) -> None:
    """
    Analyze an Excel file for potential issues.
    
    Args:
        file_path (str): Path to Excel file to analyze
    """
    if not os.path.exists(file_path):
        logger.error(f"File does not exist: {file_path}")
        return
        
    file_path = os.path.abspath(file_path)
    logger.info(f"Starting analysis of: {file_path}")
    
    # Check file size
    file_size = os.path.getsize(file_path) / (1024 * 1024)  # Size in MB
    logger.info(f"File size: {file_size:.2f} MB")
    
    # Check file permissions
    try:
        perm_info = os.stat(file_path)
        logger.info(f"File permissions: {perm_info.st_mode}")
        logger.info(f"File is readable: {os.access(file_path, os.R_OK)}")
        logger.info(f"File is writable: {os.access(file_path, os.W_OK)}")
    except Exception as e:
        logger.error(f"Error checking file permissions: {str(e)}")
    
    # Try opening with openpyxl
    try:
        logger.info("Attempting to open with openpyxl...")
        start_time = time.time()
        wb = openpyxl.load_workbook(file_path, read_only=True)
        end_time = time.time()
        
        logger.info(f"Successfully opened with openpyxl in {end_time - start_time:.2f} seconds")
        logger.info(f"Sheets in workbook: {wb.sheetnames}")
        
        # Check for specific sheet
        sheet_name = "SF132 to SF133 Reconciliation"
        if sheet_name in wb.sheetnames:
            logger.info(f"Sheet '{sheet_name}' found")
            
            # Check sheet properties
            sheet = wb[sheet_name]
            logger.info(f"Sheet dimensions: {sheet.dimensions}")
        else:
            logger.warning(f"Sheet '{sheet_name}' not found")
            
        wb.close()
    except Exception as e:
        logger.error(f"Error opening with openpyxl: {str(e)}")
    
    # Try opening with COM automation
    try:
        logger.info("Attempting to open with COM automation...")
        
        # Terminate any existing Excel processes
        for proc in psutil.process_iter(['name']):
            if proc.info['name'] and 'EXCEL.EXE' in proc.info['name'].upper():
                logger.info(f"Terminating existing Excel process: {proc.pid}")
                try:
                    proc.terminate()
                    proc.wait(timeout=5)
                except Exception as e:
                    logger.error(f"Failed to terminate Excel process: {e}")
        
        # Initialize COM
        pythoncom.CoInitialize()
        
        # Create Excel application
        start_time = time.time()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Try opening workbook
        try:
            wb = excel.Workbooks.Open(file_path)
            end_time = time.time()
            
            logger.info(f"Successfully opened with COM in {end_time - start_time:.2f} seconds")
            logger.info(f"Sheets in workbook: {[sheet.Name for sheet in wb.Sheets]}")
            
            # Check for specific sheet
            try:
                sheet = wb.Sheets("SF132 to SF133 Reconciliation")
                logger.info(f"Sheet 'SF132 to SF133 Reconciliation' found and accessed")
            except Exception as e:
                logger.warning(f"Error accessing sheet: {str(e)}")
                
            # Close workbook
            wb.Close(SaveChanges=False)
            
        except pywintypes.com_error as e:
            logger.error(f"COM error opening file: {str(e)}")
            logger.error(f"Error details: {e.args}")
            
        except Exception as e:
            logger.error(f"Error using COM automation: {str(e)}")
            
        finally:
            # Clean up COM objects
            try:
                excel.Quit()
                del excel
            except:
                pass
            
            # Force cleanup
            time.sleep(1)
            pythoncom.CoUninitialize()
            
    except Exception as e:
        logger.error(f"Error with COM initialization: {str(e)}")
    
    logger.info("Analysis complete")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        analyze_excel_file(file_path)
    else:
        file_path = input("Enter path to Excel file to analyze: ")
        analyze_excel_file(file_path)
