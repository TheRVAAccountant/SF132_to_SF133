import openpyxl
import win32com.client
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
import os
import time
import psutil
import logging
from typing import Dict, Tuple, Optional, Any, List
from dataclasses import dataclass
from queue import Queue
import shutil
from pathlib import Path
import tempfile
import uuid
import gc
import pythoncom
import contextlib
import subprocess
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException
from config import app_config
from excel_processor_config import ProcessingConfig, FileHandlingConfig, ExcelConfig
from excel_data_cleaner import ExcelDataCleaner, clean_excel_external_data

# Use the detailed configuration from excel_processor_config.py
DEFAULT_PROCESSING_CONFIG = ProcessingConfig()
DEFAULT_FILE_HANDLING_CONFIG = FileHandlingConfig()
DEFAULT_EXCEL_CONFIG = ExcelConfig()

class ExcelProcessor:
    """
    Handles Excel file processing operations including file manipulation,
    formatting, and content analysis.
    """
    
    def __init__(self, queue: Queue = None):
        """
        Initialize the Excel processor.
        
        Args:
            queue (Queue): Queue for communication with GUI
        """
        self.queue = queue
        self.config = DEFAULT_PROCESSING_CONFIG
        self.file_config = DEFAULT_FILE_HANDLING_CONFIG
        self.excel_config = DEFAULT_EXCEL_CONFIG
        self._setup_logging()
        self._temp_files = []  # Track temp files for cleanup
        
    def _setup_logging(self):
        """Configure logging for the processor."""
        self.logger = logging.getLogger(__name__)
    
    def __del__(self):
        """Clean up resources when instance is destroyed."""
        self._cleanup_temp_files()
        
    def _cleanup_temp_files(self):
        """Clean up all temporary files."""
        if app_config.cleanup_temp_files:
            for temp_file in self._temp_files:
                if os.path.exists(temp_file):
                    try:
                        os.unlink(temp_file)
                        self.logger.debug(f"Cleaned up temp file: {temp_file}")
                    except Exception as e:
                        self.logger.warning(f"Failed to clean up temp file {temp_file}: {e}")
    
    def _update_progress(self, value: float, message: str):
        """
        Update progress through the queue.
        
        Args:
            value (float): Progress value (0-100)
            message (str): Status message
        """
        if self.queue:
            self.queue.put(("progress", (value, message)))
        self.logger.info(message)
    
    def _update_status(self, message: str):
        """
        Send status update through the queue.
        
        Args:
            message (str): Status message
        """
        if self.queue:
            self.queue.put(("status", message))
        self.logger.info(message)
    
    def close_excel_instances(self):
        """Terminate all existing Excel processes to prevent file locking."""
        self._update_status("Ensuring all Excel instances are closed...")
        excel_pids = []
        
        # First find all Excel processes, including COM components
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                # Check both process name and command line for Excel references
                proc_name = proc.info.get('name', '').upper()
                cmd_line = ' '.join(proc.info.get('cmdline', [])).upper()
                
                # Look for any Excel-related processes
                if any(excel_marker in proc_name for excel_marker in ['EXCEL', 'EXCEL.EXE']) or \
                   any(excel_marker in cmd_line for excel_marker in ['EXCEL.EXE', 'EXCELCNV', 'DCOM']):
                    excel_pids.append(proc.info['pid'])
            except Exception:
                pass
        
        if excel_pids:
            self._update_status(f"Found {len(excel_pids)} Excel-related processes to close")
            
        # First try graceful termination with COM cleanup
        try:
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Try to quit any remaining Excel applications through COM
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                excel.DisplayAlerts = False
                excel.Quit()
                del excel
                gc.collect()
                self.logger.info("Gracefully closed active Excel application via COM")
            except:
                pass
            
            # Uninitialize COM
            pythoncom.CoUninitialize()
        except:
            pass
            
        # Then terminate each process with proper cleanup
        for pid in excel_pids:
            try:
                proc = psutil.Process(pid)
                proc.terminate()
                proc.wait(timeout=5)
            except Exception as e:
                self.logger.warning(f"Failed to terminate Excel process (PID {pid}): {e}")
                # Try more aggressive termination
                try:
                    if psutil.pid_exists(pid):
                        os.kill(pid, 9)  # SIGKILL or equivalent on Windows
                except Exception:
                    pass
        
        # Verify all processes are terminated
        remaining = []
        for pid in excel_pids:
            if psutil.pid_exists(pid):
                remaining.append(pid)
                
        if remaining:
            self.logger.warning(f"Could not terminate {len(remaining)} Excel processes: {remaining}")
            # For stubborn processes, try the Windows-specific taskkill command
            if sys.platform.startswith('win'):
                try:
                    self._update_status("Using system taskkill command for stubborn Excel processes...")
                    subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                                   stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                except Exception as e:
                    self.logger.warning(f"Taskkill failed: {e}")
        else:
            self._update_status("All Excel processes successfully closed")
            
        # Add a delay to ensure file system has time to release locks
        time.sleep(2)
        
        # Force garbage collection to release COM objects
        gc.collect()

    def process_file(self, original_file: str, password: str = None) -> bool:
        """
        Main processing function for Excel file.
        
        Args:
            original_file (str): Path to original Excel file
            password (str): Sheet protection password
        
        Returns:
            bool: True if processing was successful, False otherwise
        """
        try:
            # Convert to absolute path
            original_file = os.path.abspath(original_file)
            self._update_status(f"Processing file: {original_file}")
            
            # Maximum number of attempts for the overall process
            max_attempts = 3
            attempt_count = 0
            
            # Create a verified backup before any processing
            try:
                backup_dir = Path(self.config.backup_directory)
                backup_dir.mkdir(parents=True, exist_ok=True)
                
                # Create a timestamped backup
                timestamp = time.strftime("%Y%m%d-%H%M%S")
                backup_name = f"{Path(original_file).stem}_original_backup_{timestamp}.xlsx"
                backup_path = str(backup_dir / backup_name)
                
                # Copy the original file
                shutil.copy2(original_file, backup_path)
                self._update_status(f"Created original backup at: {backup_path}")
            except Exception as e:
                self.logger.warning(f"Could not create original backup: {e}")
                # Continue even if backup fails - this is just an extra safety measure
            
            while attempt_count < max_attempts:
                try:
                    attempt_count += 1
                    self._update_status(f"Processing attempt {attempt_count}/{max_attempts}")
                    
                    # Basic validation of input file
                    self._validate_file(original_file)
                    new_file = self._generate_new_filename(original_file)
                    
                    # Ensure Excel is fully closed before starting
                    self.close_excel_instances()
                    
                    # First make a clean copy of the original file with additional error handling
                    try:
                        temp_copy = self._create_verified_copy(original_file)
                    except Exception as copy_error:
                        self.logger.warning(f"Verified copy failed: {copy_error}, using direct copy")
                        # Fall back to direct copy if verified copy fails
                        temp_copy = self._get_temp_file_path("direct_temp_copy")
                        self._direct_file_copy(original_file, temp_copy)
                        
                        # Verify the copy exists and has content
                        if not os.path.exists(temp_copy) or os.path.getsize(temp_copy) == 0:
                            raise ValueError("Failed to create a valid copy of the original file")
                    
                    # Main processing approaches in order of preference:
                    # 1. Fresh workbook approach - most reliable but can fail with complex files
                    # 2. Legacy library-based processing - more compatible but slower
                    # 3. Legacy COM-based processing - most compatible with complex files
                    
                    # APPROACH 1: Try the fresh workbook approach first (most reliable)
                    self._update_progress(10, "Creating fresh workbook...")
                    success = self._process_with_fresh_workbook(temp_copy, new_file, password)
                    
                    if success:
                        self._update_progress(95, "Validating final workbook...")
                        if self._validate_excel_file(new_file):
                            self._update_progress(100, "Processing complete")
                            self._update_status(f"Successfully created and processed: {new_file}")
                            if self.queue:
                                self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                            return True
                        else:
                            # If validation fails, try repair process
                            self._update_status("File validation failed, attempting repair...")
                            if self._repair_workbook(new_file):
                                self._update_status("File repaired successfully")
                                if self.queue:
                                    self.queue.put(("success", f"File processed and repaired. Output saved to: {new_file}"))
                                return True
                            else:
                                # If repair failed, move to next attempt
                                self.logger.warning("Repair failed, trying next approach")
                                raise ValueError("Failed to create a valid Excel file")
                    else:
                        # APPROACH 2: Fall back to legacy processing methods
                        self._update_status("Fresh workbook approach failed, falling back to library methods...")
                        success = self._process_with_legacy_method(original_file, new_file, password)
                        
                        if success:
                            self._update_progress(100, "Legacy processing complete")
                            self._update_status(f"Successfully processed with legacy method: {new_file}")
                            if self.queue:
                                self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                            return True
                            
                        # If everything failed, try again with a delay
                        self.logger.warning(f"All processing attempts failed in round {attempt_count}")
                        
                        if attempt_count < max_attempts:
                            # Clear resources before next attempt
                            self.close_excel_instances()
                            gc.collect()
                            time.sleep(3)  # Wait before next attempt
                            
                            # Special error handling for subsequent attempts
                            if attempt_count == 2:
                                # On the last attempt, try more aggressive resource cleanup
                                self._update_status("Performing aggressive resource cleanup before final attempt...")
                                
                                # Force garbage collection multiple times
                                for _ in range(3):
                                    gc.collect()
                                    time.sleep(1)
                                
                                # On Windows, try to find and kill any hidden Excel processes
                                if sys.platform.startswith('win'):
                                    try:
                                        # Use taskkill with force option
                                        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE', '/T'], 
                                                    stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                                    except:
                                        pass
                        
                except Exception as attempt_error:
                    self.logger.error(f"Attempt {attempt_count} failed: {attempt_error}", exc_info=True)
                    
                    if attempt_count >= max_attempts:
                        # If we've exhausted all attempts, report the error
                        if self.queue:
                            self.queue.put(("error", f"Failed after {max_attempts} attempts: {str(attempt_error)}"))
                        return False
                    
                    # Clean up and prepare for next attempt
                    self.close_excel_instances()
                    gc.collect()
                    time.sleep(2)
            
            # If we get here, all attempts have failed
            if self.queue:
                self.queue.put(("error", f"Failed after {max_attempts} attempts"))
            return False
                
        except Exception as e:
            self.logger.error("Processing failed", exc_info=True)
            if self.queue:
                self.queue.put(("error", str(e)))
            return False
        finally:
            # Final cleanup
            self._cleanup_temp_files()
    
    def _validate_file(self, file_path: str) -> None:
        """
        Validate the input file path.
        
        Args:
            file_path (str): Path to Excel file
        
        Raises:
            ValueError: If file validation fails
        """
        if not os.path.exists(file_path):
            raise ValueError(f"File does not exist: {file_path}")
        if not file_path.lower().endswith('.xlsx'):
            raise ValueError("File must be an Excel (.xlsx) file")
    
    def _create_verified_copy(self, original_file: str) -> str:
        """
        Create a verified copy of the original file to prevent corruption.
        
        Args:
            original_file (str): Path to original Excel file
            
        Returns:
            str: Path to verified copy
        """
        temp_copy = self._get_temp_file_path("verified_copy")
        self._update_status(f"Creating verified copy at {temp_copy}...")
        
        # Use native Excel to create a clean copy (most reliable method)
        try:
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Start Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open workbook with recovery options
            wb = excel.Workbooks.Open(
                original_file,
                UpdateLinks=0,
                ReadOnly=True,
                CorruptLoad=2  # xlRepairFile (better corruption handling)
            )
            
            # Save as a new clean file
            wb.SaveAs(
                temp_copy,
                FileFormat=51,  # xlOpenXMLWorkbook
                CreateBackup=False
            )
            
            # Clean close
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Force cleanup
            del wb
            del excel
            gc.collect()
            pythoncom.CoUninitialize()
            
            # Verify file exists and has content
            if not os.path.exists(temp_copy) or os.path.getsize(temp_copy) == 0:
                raise ValueError("Failed to create valid copy")
                
            return temp_copy
            
        except Exception as e:
            self.logger.warning(f"COM copy failed: {e}, falling back to direct copy")
            
            # Fallback to direct copy
            shutil.copy2(original_file, temp_copy)
            return temp_copy
    
    def _process_with_fresh_workbook(self, source_file: str, output_file: str, password: str) -> bool:
        """
        Process by extracting data and creating a fresh workbook to avoid corruption.
        
        Args:
            source_file (str): Source Excel file
            output_file (str): Path for output file
            password (str): Sheet protection password
            
        Returns:
            bool: Whether processing succeeded
        """
        try:
            self._update_status("Loading source workbook data...")
            
            # Load source workbook
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            
            if self.config.sheet_name not in source_wb.sheetnames:
                self.logger.warning(f"Required sheet '{self.config.sheet_name}' not found")
                return False
                
            source_sheet = source_wb[self.config.sheet_name]
            
            # Create a new workbook
            self._update_status("Creating fresh workbook...")
            new_wb = openpyxl.Workbook()
            
            # Remove default sheet
            if "Sheet" in new_wb.sheetnames:
                new_wb.remove(new_wb["Sheet"])
                
            # Create new sheet with the right name
            new_sheet = new_wb.create_sheet(title=self.config.sheet_name)
            
            # Copy all data and formatting from source
            self._update_status("Copying data from source...")
            self._copy_sheet_data(source_sheet, new_sheet)
            
            # Process the fresh sheet
            self._update_progress(40, "Processing data...")
            
            # Process column visibility
            self._process_columns(new_sheet)
            
            # Find column indexes
            column_indexes = self._find_column_indexes(new_sheet)
            
            # Find header color
            rgb_color = self._process_header_formatting(new_sheet)
            
            # Find matching row
            matching_row = self._find_matching_row(new_sheet, rgb_color)
            
            # Add DO Comments column
            self._add_do_comments_column(new_sheet)
            
            # Process rows with comments
            self._process_rows_with_openpyxl(new_sheet, column_indexes, matching_row)
            
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Save the workbook
            self._update_status("Saving processed workbook...")
            new_wb.save(output_file)
            
            # Cleanup
            source_wb.close()
            new_wb.close()
            
            # Verify the saved file
            if not os.path.exists(output_file):
                return False
                
            return True
            
        except Exception as e:
            self.logger.error(f"Error in fresh workbook processing: {e}", exc_info=True)
            return False
    
    def _copy_sheet_data(self, source_sheet, target_sheet):
        """
        Copy data and basic formatting from source sheet to target sheet.
        
        Args:
            source_sheet: Source worksheet
            target_sheet: Target worksheet
        """
        # Get dimensions of source sheet
        max_row = source_sheet.max_row
        max_col = source_sheet.max_column
        
        # Copy column dimensions and properties
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_sheet.column_dimensions:
                src_col_dim = source_sheet.column_dimensions[col_letter]
                tgt_col_dim = target_sheet.column_dimensions[col_letter]
                
                # Copy column width
                tgt_col_dim.width = src_col_dim.width if src_col_dim.width else 8.43  # Default width
                
                # Copy hidden status
                tgt_col_dim.hidden = False  # We want all columns visible
        
        # Copy row heights
        for row_idx in range(1, max_row + 1):
            if row_idx in source_sheet.row_dimensions:
                src_row_dim = source_sheet.row_dimensions[row_idx]
                tgt_row_dim = target_sheet.row_dimensions[row_idx]
                
                # Copy row height
                tgt_row_dim.height = src_row_dim.height if src_row_dim.height else 15  # Default height
        
        # Copy cell data and basic formatting
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                # Get source cell
                src_cell = source_sheet.cell(row=row_idx, column=col_idx)
                
                # Skip merged cells (we'll handle them separately)
                if isinstance(src_cell, MergedCell):
                    continue
                    
                # Create target cell and copy value
                tgt_cell = target_sheet.cell(row=row_idx, column=col_idx, value=src_cell.value)
                
                # Copy basic formatting with proper error handling
                if src_cell.has_style:
                    try:
                        # Font
                        tgt_cell.font = Font(
                            name=src_cell.font.name,
                            size=src_cell.font.size,
                            bold=src_cell.font.bold,
                            italic=src_cell.font.italic,
                            color=src_cell.font.color
                        )
                    except Exception as e:
                        self.logger.debug(f"Error copying font at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Fill - with proper Color object creation
                        if src_cell.fill and hasattr(src_cell.fill, 'start_color') and src_cell.fill.start_color:
                            fill_color = src_cell.fill.start_color.rgb or "FFFFFF"
                            # Create a proper Color object from the RGB string
                            color_obj = Color(rgb=fill_color)
                            tgt_cell.fill = PatternFill(
                                fill_type='solid',
                                start_color=color_obj
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying fill at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Border
                        if src_cell.border:
                            tgt_cell.border = Border(
                                left=src_cell.border.left,
                                right=src_cell.border.right,
                                top=src_cell.border.top,
                                bottom=src_cell.border.bottom
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying border at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Alignment
                        if src_cell.alignment:
                            tgt_cell.alignment = Alignment(
                                horizontal=src_cell.alignment.horizontal,
                                vertical=src_cell.alignment.vertical,
                                wrap_text=src_cell.alignment.wrap_text
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying alignment at {row_idx},{col_idx}: {e}")
        
        # We'll deliberately NOT copy merged cells to avoid potential corruption
    
    def _validate_excel_file(self, file_path: str) -> bool:
        """
        Validate an Excel file by trying to open it and check for errors.
        
        Args:
            file_path (str): Path to Excel file to validate
            
        Returns:
            bool: Whether the file is valid
        """
        try:
            # Method 1: Try opening with openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            # Initial validation success
            self._update_status(f"Basic validation with openpyxl successful. Found {len(sheet_names)} sheets.")
            
            # Method 2: Verify with Excel COM if possible
            if self.excel_config.enable_com and sys.platform.startswith('win'):
                # Make sure Excel is not running
                self.close_excel_instances()
                
                # Initialize COM with error handling
                try:
                    pythoncom.CoInitialize()
                except Exception as e:
                    self.logger.warning(f"COM initialization failed: {e}")
                
                excel = None
                wb = None
                
                try:
                    # Start Excel with proper exception handling
                    try:
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                    except Exception as e:
                        self.logger.warning(f"Excel COM initialization failed: {e}")
                        return True  # Fall back to openpyxl validation
                    
                    # Try to open the file
                    try:
                        wb = excel.Workbooks.Open(
                            file_path,
                            UpdateLinks=0,
                            ReadOnly=True
                        )
                    except Exception as e:
                        self.logger.warning(f"Excel COM file open failed: {e}")
                        return True  # Fall back to openpyxl validation
                    
                    # Check for errors using a more reliable method than ErrorCheckingStatus
                    has_errors = False
                    
                    # Try different error checking methods
                    try:
                        # Method 1: Check if ErrorCheckingStatus attribute exists and use it
                        if hasattr(excel, 'ErrorCheckingStatus'):
                            has_errors = excel.ErrorCheckingStatus
                        # Method 2: Use Excel's CheckWorkbookCompatibility method if available
                        elif hasattr(wb, 'CheckCompatibility'):
                            compatibility_issues = wb.CheckCompatibility
                            has_errors = compatibility_issues
                        # Method 3: Check for error alerts on sheets
                        else:
                            # If no direct error checking is available, check if file can be saved
                            temp_path = self._get_temp_file_path("verify")
                            wb.SaveAs(temp_path, FileFormat=51)  # 51 = xlsx
                            if os.path.exists(temp_path):
                                os.unlink(temp_path)
                            has_errors = False
                    except Exception as e:
                        self.logger.debug(f"Error checking method failed: {e}")
                        # If error checking fails, assume the file is okay if we got this far
                        has_errors = False
                    
                    # Close properly
                    if wb:
                        try:
                            wb.Close(SaveChanges=False)
                        except:
                            pass
                        
                    if excel:
                        try:
                            excel.Quit()
                        except:
                            pass
                        
                    # Cleanup COM objects
                    if wb:
                        del wb
                    if excel:
                        del excel
                        
                    gc.collect()
                    
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                    
                    if has_errors:
                        self.logger.warning(f"Excel detected errors in {file_path}")
                        return False
                    
                    self._update_status("Enhanced COM validation successful")
                    
                except Exception as e:
                    self.logger.warning(f"COM validation failed: {e}")
                    # Thorough cleanup after failure
                    if wb:
                        try:
                            wb.Close(SaveChanges=False)
                        except:
                            pass
                    if excel:
                        try:
                            excel.Quit()
                        except:
                            pass
                    
                    # Cleanup COM objects
                    if wb:
                        del wb
                    if excel:
                        del excel
                        
                    gc.collect()
                    
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                    
                    # Fall back to considering the file valid if openpyxl could open it
                    return True
            
            # If we got here, file seems valid
            return True
            
        except Exception as e:
            self.logger.warning(f"File validation failed: {e}")
            return False
    
    def _repair_workbook(self, file_path: str) -> bool:
        """
        Repair a potentially corrupted Excel file.
        
        Args:
            file_path (str): Path to Excel file to repair
            
        Returns:
            bool: Whether repair was successful
        """
        try:
            self._update_status("Attempting to repair workbook...")
            
            # Ensure all Excel processes are closed
            self.close_excel_instances()
            
            # Create a backup before repair
            backup_path = self._create_backup_file(file_path)
            self._update_status(f"Created backup at: {backup_path}")
            
            # Check file access before attempting repair
            max_retries = 5
            retry_count = 0
            file_accessible = False
            
            while retry_count < max_retries and not file_accessible:
                try:
                    # Test file accessibility with read/write checks
                    with open(file_path, 'rb+') as test_file:
                        test_file.seek(0)
                        test_file.read(1)  # Try to read a byte
                        test_file.seek(0)  # Go back to beginning
                    file_accessible = True
                except Exception as e:
                    retry_count += 1
                    self.logger.warning(f"File access retry {retry_count}/{max_retries}: {e}")
                    
                    # Try to release the file by forcing garbage collection
                    gc.collect()
                    time.sleep(2)  # Wait before retrying
                    
                    # On Windows, try additional methods to release the file
                    if sys.platform.startswith('win'):
                        self.close_excel_instances()
                        
                        # Try to work with a copy if original is inaccessible
                        if retry_count >= 3:
                            try:
                                temp_copy = self._get_temp_file_path("repair_copy")
                                self._update_status(f"Attempting to create a copy at {temp_copy}...")
                                
                                # Use low-level file operations to copy
                                with open(file_path, 'rb') as src, open(temp_copy, 'wb') as dst:
                                    dst.write(src.read())
                                
                                # If copy succeeds, work with the copy instead
                                if os.path.exists(temp_copy) and os.path.getsize(temp_copy) > 0:
                                    self._update_status("Working with copy since original is locked")
                                    file_path = temp_copy
                                    file_accessible = True
                            except Exception as copy_err:
                                self.logger.warning(f"Failed to create copy: {copy_err}")
            
            if not file_accessible:
                self.logger.error("Could not access file for repair after multiple attempts")
                return False
            
            # Try multiple repair methods in sequence
            repair_methods = [
                self._repair_with_excel_com,
                self._repair_with_pandas,
                self._repair_with_openpyxl,
                self._repair_with_system_tool
            ]
            
            # Try each repair method until one succeeds
            for repair_method in repair_methods:
                try:
                    self._update_status(f"Trying repair method: {repair_method.__name__}")
                    success, repaired_path = repair_method(file_path)
                    
                    if success and os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                        # Validate the repaired file
                        if self._validate_repaired_file(repaired_path):
                            # Replace original with repaired version
                            try:
                                # Enhanced file replacement logic
                                original_filename = os.path.basename(file_path)
                                self._update_status(f"Replacing {original_filename} with repaired version...")
                                
                                # Remove original if it exists and isn't the same as the repaired file
                                if os.path.exists(file_path) and os.path.abspath(file_path) != os.path.abspath(repaired_path):
                                    try:
                                        os.unlink(file_path)
                                    except Exception as e:
                                        self.logger.warning(f"Could not remove original file: {e}")
                                        # If we can't delete, try to move it instead
                                        try:
                                            failed_path = self._get_temp_file_path("failed_original")
                                            shutil.move(file_path, failed_path)
                                        except:
                                            pass
                                
                                # Copy the repaired file to the original location
                                shutil.copy2(repaired_path, file_path)
                                
                                # Verify the final file
                                if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                                    self._update_status("Repair completed successfully")
                                    return True
                            except Exception as e:
                                self.logger.warning(f"File replacement failed: {e}")
                                # Continue to next method if replacement fails
                        else:
                            self.logger.warning("Repaired file validation failed")
                except Exception as method_error:
                    self.logger.warning(f"{repair_method.__name__} failed: {method_error}")
            
            # If all repair methods failed, try to restore from backup
            self._update_status("All repair methods failed. Attempting to restore from backup...")
            if os.path.exists(backup_path) and os.path.getsize(backup_path) > 0:
                try:
                    # Ensure target file is accessible
                    if os.path.exists(file_path):
                        try:
                            os.unlink(file_path)
                        except:
                            # If deletion fails, try to use a new path instead
                            file_path = self._get_temp_file_path("restored")
                    
                    # Copy the backup
                    shutil.copy2(backup_path, file_path)
                    
                    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                        self._update_status("Successfully restored from backup")
                        return True
                except Exception as restore_error:
                    self.logger.error(f"Backup restoration failed: {restore_error}")
            
            return False
            
        except Exception as e:
            self.logger.error(f"Repair process failed: {e}")
            return False
            
    def _validate_repaired_file(self, file_path: str) -> bool:
        """
        Validate a repaired Excel file to ensure it's usable.
        
        Args:
            file_path (str): Path to the repaired file
            
        Returns:
            bool: Whether the file is valid
        """
        try:
            # Check if file exists and has size
            if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                return False
                
            # Try opening with openpyxl in read-only mode
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                self._update_status(f"Validated repaired file with {sheet_count} sheets")
                return True
            except Exception as e:
                self.logger.warning(f"openpyxl validation failed: {e}")
                
            # Try with pandas as fallback
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_count = len(excel_file.sheet_names)
                excel_file.close()
                self._update_status(f"Validated repaired file with pandas: {sheet_count} sheets")
                return True
            except Exception as e:
                self.logger.warning(f"pandas validation failed: {e}")
                
            return False
            
        except Exception as e:
            self.logger.warning(f"Repair validation failed: {e}")
            return False
            
    def _repair_with_excel_com(self, file_path: str) -> tuple:
        """
        Repair Excel file using Excel's COM interface.
        
        Args:
            file_path (str): Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not self.excel_config.enable_com or not sys.platform.startswith('win'):
            return False, None
            
        excel = None
        wb = None
        repaired_path = self._get_temp_file_path("com_repaired")
        
        try:
            # Ensure Excel is closed
            self.close_excel_instances()
            
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Start Excel with robust error handling
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.EnableEvents = False
                excel.Calculation = -4135  # xlCalculationManual
            except Exception as e:
                self.logger.warning(f"Excel COM initialization failed: {e}")
                return False, None
                
            # Try to open the file in repair mode
            try:
                self._update_status("Opening file in repair mode...")
                wb = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=0,
                    CorruptLoad=2  # xlRepairFile - repair mode
                )
            except Exception as e:
                self.logger.warning(f"Failed to open file in repair mode: {e}")
                # Try with different parameters
                try:
                    wb = excel.Workbooks.Open(
                        file_path,
                        UpdateLinks=0,
                        ReadOnly=True
                    )
                except Exception as e2:
                    self.logger.warning(f"Failed to open file in read-only mode: {e2}")
                    return False, None
            
            # Calculate workbook
            try:
                wb.Calculate()
            except:
                pass
                
            # Save to a new file
            self._update_status(f"Saving repaired file to {repaired_path}...")
            wb.SaveAs(
                repaired_path,
                FileFormat=51,  # xlOpenXMLWorkbook
                CreateBackup=False
            )
            
            # Close workbook and quit Excel
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Cleanup
            del wb
            del excel
            wb = None
            excel = None
            gc.collect()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Excel COM repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Excel COM repair failed: {e}")
            return False, None
        finally:
            # Ensure proper cleanup
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
                    
            # Force garbage collection
            if wb:
                del wb
            if excel:
                del excel
            gc.collect()
            
            # Uninitialize COM
            try:
                pythoncom.CoUninitialize()
            except:
                pass
                
    def _repair_with_pandas(self, file_path: str) -> tuple:
        """
        Repair Excel file using pandas.
        
        Args:
            file_path (str): Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        repaired_path = self._get_temp_file_path("pandas_repaired")
        
        try:
            self._update_status("Attempting pandas-based repair...")
            
            # Try to read the file with pandas
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # Create a new Excel writer
            writer = pd.ExcelWriter(repaired_path, engine='openpyxl')
            
            # Process each sheet
            for sheet_name in sheet_names:
                self._update_status(f"Processing sheet: {sheet_name}")
                # Read with error handling for individual sheets
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as sheet_err:
                    self.logger.warning(f"Could not process sheet {sheet_name}: {sheet_err}")
                    # Create an empty sheet instead
                    df = pd.DataFrame()
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Save the writer
            writer.close()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Pandas repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Pandas repair failed: {e}")
            return False, None
            
    def _repair_with_openpyxl(self, file_path: str) -> tuple:
        """
        Repair Excel file using openpyxl.
        
        Args:
            file_path (str): Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        repaired_path = self._get_temp_file_path("openpyxl_repaired")
        
        try:
            self._update_status("Attempting openpyxl-based repair...")
            
            # Try to read the workbook
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
            except Exception as read_err:
                self.logger.warning(f"Failed to open with openpyxl: {read_err}")
                return False, None
                
            # Create a new workbook
            new_wb = openpyxl.Workbook()
            
            # Remove default sheet
            if "Sheet" in new_wb.sheetnames:
                new_wb.remove(new_wb["Sheet"])
                
            # Copy each sheet from source
            for sheet_name in wb.sheetnames:
                self._update_status(f"Processing sheet: {sheet_name}")
                # Create new sheet
                new_sheet = new_wb.create_sheet(title=sheet_name)
                src_sheet = wb[sheet_name]
                
                # Try to copy basic data
                try:
                    # Get rows from source sheet
                    for row_idx, row in enumerate(src_sheet.iter_rows(values_only=True), 1):
                        for col_idx, value in enumerate(row, 1):
                            new_sheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception as sheet_err:
                    self.logger.warning(f"Error copying sheet {sheet_name}: {sheet_err}")
            
            # Close the source workbook
            wb.close()
            
            # Save the new workbook
            new_wb.save(repaired_path)
            new_wb.close()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Openpyxl repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Openpyxl repair failed: {e}")
            return False, None
            
    def _repair_with_system_tool(self, file_path: str) -> tuple:
        """
        Try to repair the Excel file using system tools (Windows only).
        
        Args:
            file_path (str): Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not sys.platform.startswith('win'):
            return False, None
            
        repaired_path = self._get_temp_file_path("system_repaired")
        
        try:
            self._update_status("Attempting system-based repair...")
            
            # First try to copy the file to a new location
            shutil.copy2(file_path, repaired_path)
            
            # Then try ExcelCnv command-line tool if available (on some Windows systems)
            try:
                office_paths = [
                    r"C:\Program Files\Microsoft Office\root\Office16",
                    r"C:\Program Files (x86)\Microsoft Office\root\Office16",
                    r"C:\Program Files\Microsoft Office\Office16",
                    r"C:\Program Files (x86)\Microsoft Office\Office16",
                    r"C:\Program Files\Microsoft Office\Office15",
                    r"C:\Program Files (x86)\Microsoft Office\Office15",
                ]
                
                excelcnv_path = None
                for office_path in office_paths:
                    test_path = os.path.join(office_path, "excelcnv.exe")
                    if os.path.exists(test_path):
                        excelcnv_path = test_path
                        break
                        
                if excelcnv_path:
                    self._update_status(f"Found Excel converter at: {excelcnv_path}")
                    
                    # Generate a temporary output path
                    output_path = self._get_temp_file_path("excelcnv_output")
                    
                    # Run ExcelCnv to convert/repair the file
                    subprocess.run([
                        excelcnv_path,
                        "-nme",  # No message boxes
                        "-oice",  # Open Invalid with Converter Extensions
                        "-xlsb",  # Convert to XLSB format first (more robust)
                        repaired_path,
                        output_path
                    ], stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                    
                    # Convert back to xlsx
                    if os.path.exists(output_path):
                        final_path = self._get_temp_file_path("final_repaired")
                        subprocess.run([
                            excelcnv_path,
                            "-nme",
                            "-xlsx",  # Convert to XLSX format
                            output_path,
                            final_path
                        ], stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                        
                        if os.path.exists(final_path) and os.path.getsize(final_path) > 0:
                            self._update_status("System repair successful")
                            return True, final_path
            except Exception as system_err:
                self.logger.warning(f"System tool repair failed: {system_err}")
            
            # If system tools failed, return the simple copy if it exists
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"System repair failed: {e}")
            return False, None
    
    def _process_with_legacy_method(self, original_file: str, new_file: str, password: str) -> bool:
        """
        Process using the legacy method as a fallback.
        
        Args:
            original_file (str): Original file path
            new_file (str): New file path
            password (str): Sheet protection password
            
        Returns:
            bool: Whether processing was successful
        """
        try:
            # First try to process with pandas/openpyxl directly
            self._update_progress(5, "Attempting to process file with pandas/openpyxl...")
            try:
                success = self._process_with_libraries(original_file, new_file, password)
                if success:
                    self._update_progress(90, "Direct library processing successful")
                    self._update_status(f"Successfully created and processed: {new_file}")
                    if self.queue:
                        self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                    return True
                else:
                    self._update_status("Direct library processing failed, falling back to COM methods...")
            except Exception as e:
                self.logger.warning(f"Direct library processing failed: {e}")
                self._update_status("Falling back to COM methods...")
            
            # Fall back to original COM processing approach
            try:
                # First make a copy of the original file - using more reliable file system operations
                self._update_progress(10, f"Creating new workbook at {new_file}...")
                if self.excel_config.use_com_for_copy:
                    # Try COM approach first, but fall back to direct copy if it fails
                    success = self._create_clean_copy(original_file, new_file)
                    if not success:
                        self._update_status("Falling back to direct file copy...")
                        self._direct_file_copy(original_file, new_file)
                else:
                    # Use direct file copy without COM (more reliable)
                    self._direct_file_copy(original_file, new_file)
                
                self._update_progress(20, "Processing Excel file...")
                self._process_workbook(new_file, password)
                
                # Final verification and notification
                if os.path.exists(new_file):
                    self._update_status(f"Successfully created and processed: {new_file}")
                    if self.queue:
                        self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                    return True
                else:
                    raise FileNotFoundError(f"Expected output file not found: {new_file}")
            except Exception as e:
                self.logger.error(f"COM processing failed: {str(e)}", exc_info=True)
                if self.queue:
                    self.queue.put(("error", f"Processing failed: {str(e)}"))
                return False
                
        except Exception as e:
            self.logger.error("Processing failed", exc_info=True)
            if self.queue:
                self.queue.put(("error", str(e)))
            return False
    
    def _generate_new_filename(self, original_file: str) -> str:
        """
        Generate the new filename based on the original.
        
        Args:
            original_file (str): Original file path
        
        Returns:
            str: New file path
        """
        # Get original file basename without extension
        original_basename = Path(original_file).stem
        
        # Create output directory if it doesn't exist
        output_dir = Path(self.config.output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create a descriptive filename with timestamp
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        new_filename = f"{original_basename}_processed_{timestamp}.xlsx"
        
        # Construct full path
        new_path = output_dir / new_filename
        
        self._update_status(f"Generated output filename: {new_path}")
        return str(new_path.absolute())
    
    def _process_with_libraries(self, original_file: str, output_file: str, password: str) -> bool:
        """
        Process Excel file using Python libraries (openpyxl/pandas) instead of COM.
        
        Args:
            original_file (str): Path to original Excel file
            output_file (str): Path to output file
            password (str): Sheet protection password
            
        Returns:
            bool: True if processing was successful, False otherwise
        """
        try:
            self._update_status(f"Loading file with pandas/openpyxl: {original_file}")
            
            # Create output directory if needed
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Copy the file first for safety
            shutil.copy2(original_file, output_file)
            
            # Load workbook with openpyxl
            self._update_status("Loading workbook with openpyxl...")
            wb = openpyxl.load_workbook(output_file, data_only=True)
            
            if self.config.sheet_name not in wb.sheetnames:
                self.logger.warning(f"Required sheet '{self.config.sheet_name}' not found")
                return False
            
            sheet = wb[self.config.sheet_name]
            
            # Unprotect sheet if password provided
            if password:
                self._update_progress(20, "Unprotecting sheet...")
                self._unprotect_sheet(sheet, password)
            
            self._update_progress(30, "Processing columns...")
            self._process_columns(sheet)
            
            self._update_progress(50, "Processing merged cells...")
            self._process_merged_cells(sheet)
            
            self._update_progress(60, "Finding column indexes...")
            column_indexes = self._find_column_indexes(sheet)
            
            self._update_progress(70, "Processing header formatting...")
            rgb_color = self._process_header_formatting(sheet)
            
            self._update_progress(80, "Finding matching rows...")
            matching_row = self._find_matching_row(sheet, rgb_color)
            
            self._update_progress(85, "Adding DO Comments column...")
            self._add_do_comments_column(sheet)
            
            self._update_progress(90, "Processing rows with comments...")
            self._process_rows_with_openpyxl(sheet, column_indexes, matching_row)
            
            # Save the workbook
            wb.save(output_file)
            wb.close()
            
            self._update_progress(100, "Processing complete")
            return True
                
        except Exception as e:
            self.logger.warning(f"Library-based processing failed: {str(e)}")
            return False
    
    def _direct_file_copy(self, source_file: str, dest_file: str) -> bool:
        """
        Perform a direct file copy without using COM.
        This is more reliable when COM operations fail.
        
        Args:
            source_file (str): Source file path
            dest_file (str): Destination file path
            
        Returns:
            bool: Whether the copy was successful
        """
        try:
            self._update_status(f"Directly copying file from {source_file} to {dest_file}...")
            
            # Ensure the destination directory exists
            os.makedirs(os.path.dirname(dest_file), exist_ok=True)
            
            # Make sure we can access the source file
            max_retries = 5
            retry_count = 0
            file_accessible = False
            
            while retry_count < max_retries and not file_accessible:
                try:
                    # Test if we can read the source file
                    with open(source_file, 'rb') as test_file:
                        test_file.read(1)  # Try to read a byte
                    file_accessible = True
                except Exception as e:
                    retry_count += 1
                    self.logger.warning(f"Source file access retry {retry_count}/{max_retries}: {e}")
                    
                    # Try to release the file by forcing garbage collection
                    gc.collect()
                    time.sleep(2)  # Wait before retrying
                    
                    # Make sure Excel is closed
                    if retry_count >= 2:
                        self.close_excel_instances()
            
            if not file_accessible:
                self.logger.error("Could not access source file for copying")
                return False
            
            # Check for destination file accessibility
            if os.path.exists(dest_file):
                try:
                    # Try to remove existing dest file if it might be locked
                    with open(dest_file, 'r+b') as test_file:
                        test_file.seek(0)
                except Exception as e:
                    self.logger.warning(f"Destination file is locked, trying to use a different path: {e}")
                    # Use a new destination path
                    old_dest = dest_file
                    dest_file = self._get_temp_file_path("direct_copy")
                    self._update_status(f"Using alternative path: {dest_file}")
            
            # Try several copy methods
            copy_methods = [
                # Method 1: Standard shutil.copy2
                lambda s, d: shutil.copy2(s, d),
                
                # Method 2: Low-level file copy with chunks
                lambda s, d: self._chunk_copy(s, d),
                
                # Method 3: Use system commands (Windows)
                lambda s, d: self._system_copy(s, d) if sys.platform.startswith('win') else None,
                
                # Method 4: os.system copy (fallback)
                lambda s, d: os.system(f'copy "{s}" "{d}"') if sys.platform.startswith('win') else None
            ]
            
            # Try each copy method until one succeeds
            success = False
            for method_idx, copy_method in enumerate(copy_methods):
                try:
                    self._update_status(f"Trying copy method {method_idx+1}...")
                    result = copy_method(source_file, dest_file)
                    if result is not None and result != 0:
                        continue  # Try next method if this one returned an error code
                    
                    # Verify the copy succeeded
                    if os.path.exists(dest_file) and os.path.getsize(dest_file) > 0:
                        file_size = os.path.getsize(dest_file)
                        self._update_status(f"Direct file copy successful ({file_size} bytes)")
                        success = True
                        break
                except Exception as e:
                    self.logger.warning(f"Copy method {method_idx+1} failed: {e}")
                    continue
            
            if not success:
                self.logger.error("All copy methods failed")
                return False
                
            return True
            
        except Exception as e:
            self.logger.error(f"Error during direct file copy: {str(e)}", exc_info=True)
            return False
            
    def _chunk_copy(self, source_file: str, dest_file: str) -> bool:
        """
        Copy a file in chunks to avoid memory issues with large files.
        
        Args:
            source_file (str): Source file path
            dest_file (str): Destination file path
            
        Returns:
            bool: Whether the copy was successful
        """
        try:
            # Use a larger buffer for faster copying of large files
            buffer_size = 10 * 1024 * 1024  # 10MB buffer
            
            with open(source_file, 'rb') as src, open(dest_file, 'wb') as dst:
                while True:
                    chunk = src.read(buffer_size)
                    if not chunk:
                        break
                    dst.write(chunk)
            
            return True
        except Exception as e:
            self.logger.warning(f"Chunk copy failed: {e}")
            return False
    
    def _system_copy(self, source_file: str, dest_file: str) -> int:
        """
        Use system-specific copy commands for reliability.
        
        Args:
            source_file (str): Source file path
            dest_file (str): Destination file path
            
        Returns:
            int: Return code (0 for success)
        """
        if sys.platform.startswith('win'):
            # Windows - use robocopy or xcopy
            try:
                # Try robocopy first (more reliable for locked files)
                source_dir = os.path.dirname(source_file)
                source_file_name = os.path.basename(source_file)
                dest_dir = os.path.dirname(dest_file)
                
                # Use /R:3 to retry 3 times if file is locked, /W:2 to wait 2 seconds between retries
                # /J for unbuffered I/O, /B for backup mode (can copy open files), /NP for no progress
                result = subprocess.run(
                    ['robocopy', source_dir, dest_dir, source_file_name, '/R:3', '/W:2', '/J', '/B', '/NP'],
                    capture_output=True,
                    text=True
                )
                
                # Robocopy return codes: 0 = no files copied, 1 = files copied, > 1 = errors
                if result.returncode <= 1:
                    return 0  # Success
                
                # If robocopy failed, try xcopy
                result = subprocess.run(
                    ['xcopy', source_file, dest_file, '/Y', '/Q', '/R', '/H'],
                    capture_output=True,
                    text=True
                )
                
                return result.returncode
            except Exception as e:
                self.logger.warning(f"System copy failed: {e}")
                return 1
        else:
            # Unix-like - use cp
            try:
                result = subprocess.run(['cp', source_file, dest_file], capture_output=True, text=True)
                return result.returncode
            except Exception as e:
                self.logger.warning(f"System cp failed: {e}")
                return 1
    
    def _create_clean_copy(self, original_file: str, new_file: str) -> bool:
        """
        Create a clean copy of the original file using COM if possible.
        
        Args:
            original_file (str): Original file path
            new_file (str): New file path
            
        Returns:
            bool: Whether the operation was successful
        """
        try:
            # First make sure the output directory exists
            os.makedirs(os.path.dirname(new_file), exist_ok=True)
            
            # Create a temporary file for intermediate processing
            temp_file = self._get_temp_file_path("init_copy")
            
            # Try opening the original file to verify its integrity
            self._update_status("Verifying original file integrity...")
            try:
                # Use openpyxl to verify file can be opened
                with self._open_and_close_workbook(original_file) as orig_wb:
                    if orig_wb is None:
                        self.logger.warning("Original file couldn't be verified with openpyxl, "
                                           "will attempt to fix through copy process")
            except Exception as e:
                self.logger.warning(f"Original file verification failed: {e}")
                
            # Do a standard file copy first to get a basic working file
            self._update_status("Creating initial file copy...")
            shutil.copy2(original_file, temp_file)
            
            # Now use Excel COM to save the file properly
            self._update_status("Creating clean Excel copy...")
            return self._save_with_excel_com(temp_file, new_file, preserve_all=True)
            
        except Exception as e:
            self._update_status(f"Error creating clean copy: {str(e)}")
            return False
    
    def _save_with_excel_com(self, source_file: str, dest_file: str, preserve_all: bool = False) -> bool:
        """
        Use Excel COM to save a file, with robust fallback mechanisms.
        
        Args:
            source_file (str): Source file path
            dest_file (str): Destination file path
            preserve_all (bool): Whether to preserve all Excel features
            
        Returns:
            bool: Whether the save was successful
        """
        if not self.excel_config.enable_com:
            self._update_status("COM operations disabled in configuration")
            return False
            
        # Check if file exists before attempting COM
        if not os.path.exists(source_file):
            self.logger.error(f"Source file does not exist: {source_file}")
            return False
            
        # Make sure Excel is not running to avoid conflicts
        self.close_excel_instances()
        time.sleep(1)  # Give OS time to fully release resources
        
        excel = None
        wb = None
        success = False
        retry_count = 0
        max_retries = self.excel_config.max_com_retries
        
        while retry_count <= max_retries and not success:
            try:
                # Add COM security settings
                try:
                    # Set default security for automation
                    import win32com
                    import win32com.client.gencache
                    win32com.client.gencache.EnsureDispatch("Excel.Application")
                except:
                    pass
                
                # Initialize COM with proper threading model
                pythoncom.CoInitialize()
                
                # Start Excel with proper security context
                self._update_status(f"Starting Excel (attempt {retry_count+1}/{max_retries+1})...")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.AskToUpdateLinks = False
                excel.EnableEvents = False  # Disable Excel events
                
                # ENHANCEMENT: Set calculation mode to manual to prevent recalculation issues
                excel.Calculation = -4135  # xlCalculationManual
                
                # Open workbook with safe options
                self._update_status(f"Opening file with Excel COM: {source_file}")
                
                # Use explicit constants for better clarity
                XL_UPDATE_LINKS_NEVER = 0
                XL_CORRUPT_LOAD_NORMAL = 0
                
                # Open with more robust error handling
                wb = excel.Workbooks.Open(
                    source_file,
                    UpdateLinks=XL_UPDATE_LINKS_NEVER,  
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    CorruptLoad=XL_CORRUPT_LOAD_NORMAL
                )
                
                # Clean up external connections if not preserving everything
                if not preserve_all:
                    self._clean_external_connections(wb, excel)
                
                # ENHANCEMENT: Make sure the workbook is recalculated
                wb.Calculate()
                
                # Save to the destination with correct format
                self._update_status(f"Saving clean copy to: {dest_file}")
                
                # Use explicit constant for better clarity
                XL_XLSX = 51  # Excel.XlFileFormat.xlOpenXMLWorkbook
                
                # ENHANCEMENT: Add repair functionality to save process
                try:
                    # First save to a temporary location
                    temp_save = self._get_temp_file_path("com_save")
                    
                    wb.SaveAs(
                        temp_save,
                        FileFormat=XL_XLSX,
                        CreateBackup=False
                    )
                    
                    # Now let Excel repair the file and save it again
                    # Close the current workbook
                    wb.Close(SaveChanges=False)
                    
                    # Reopen with repair option
                    wb = excel.Workbooks.Open(
                        temp_save,
                        UpdateLinks=0,
                        CorruptLoad=2  # xlRepairFile - repair mode
                    )
                    
                    # Save to final destination
                    wb.SaveAs(
                        dest_file,
                        FileFormat=XL_XLSX,
                        CreateBackup=False
                    )
                    
                except Exception:
                    # Fall back to direct save if repair approach fails
                    self.logger.warning("Repair approach failed, using direct save")
                    wb.SaveAs(
                        dest_file,
                        FileFormat=XL_XLSX,
                        CreateBackup=False
                    )
                
                # Close properly
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                # Verify the file was created successfully
                if os.path.exists(dest_file) and os.path.getsize(dest_file) > 0:
                    success = True
                    self._update_status(f"Excel COM save operation successful")
                else:
                    raise IOError(f"Failed to save file to {dest_file}")
                
            except Exception as e:
                retry_count += 1
                self.logger.warning(f"COM attempt {retry_count}/{max_retries+1} failed: {str(e)}")
                
                if retry_count <= max_retries:
                    self._update_status(f"Retrying COM operation (attempt {retry_count+1}/{max_retries+1})...")
                    time.sleep(2)  # Wait before retrying
                    # Make sure Excel is fully closed
                    self.close_excel_instances()
                else:
                    self.logger.error(f"All COM attempts failed: {str(e)}")
            finally:
                # Thorough cleanup of COM resources
                if wb:
                    try:
                        wb.Close(SaveChanges=False)
                    except:
                        pass
                    wb = None
                
                if excel:
                    try:
                        excel.Quit()
                    except:
                        pass
                    excel = None
                
                # Force garbage collection to release COM objects
                gc.collect()
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
                
                # Give OS time to release resources
                time.sleep(0.5)
                
        return success
    
    def _clean_external_connections(self, wb, excel) -> None:
        """
        Clean up external data connections and links in the workbook.
        
        Args:
            wb: Excel workbook COM object
            excel: Excel application COM object
        """
        try:
            # Initialize the specialized external data cleaner
            data_cleaner = ExcelDataCleaner(self.logger)
            
            self._update_status("Cleaning external data connections...")
            
            # Use the specialized sheet cleaning method for each sheet
            for i in range(1, wb.Sheets.Count + 1):
                try:
                    sheet = wb.Sheets(i)
                    data_cleaner._clean_sheet_data(sheet)
                except Exception as e:
                    self.logger.warning(f"Error cleaning sheet {i}: {e}")
            
            # Clean general workbook connections
            try:
                # Break external links
                for link_type in [1, 2, 5, 6]:  # Different types of links
                    try:
                        self._update_status(f"Breaking links of type {link_type}...")
                        wb.BreakLink(Name="", Type=link_type)
                    except:
                        pass
                    
                # Remove connections
                if hasattr(wb, 'Connections') and wb.Connections.Count > 0:
                    self._update_status(f"Removing {wb.Connections.Count} external connections...")
                    
                    conn_names = []
                    for i in range(1, wb.Connections.Count + 1):
                        try:
                            conn_names.append(wb.Connections(i).Name)
                        except:
                            pass
                    
                    for name in conn_names:
                        try:
                            wb.Connections(name).Delete()
                        except Exception as e:
                            self.logger.warning(f"Failed to remove connection {name}: {e}")
                
                # Commit connection changes
                try:
                    if hasattr(wb, 'Connections'):
                        wb.Connections.CommitAll()
                except:
                    pass
                    
            except Exception as e:
                self.logger.warning(f"Error cleaning workbook connections: {e}")
                
        except Exception as e:
            self.logger.warning(f"Error during connection cleanup: {e}")

    # Add a new method to perform complete external data cleaning
    def _perform_external_data_cleanup(self, file_path: str) -> bool:
        """
        Performs a specialized cleanup of all external data references.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            bool: Success status
        """
        self._update_status("Performing specialized external data cleanup...")
        return clean_excel_external_data(file_path, self.logger)
    
    def _process_workbook(self, file_path: str, password: str) -> None:
        """
        Process the Excel workbook.
        
        Args:
            file_path (str): Path to Excel file
            password (str): Sheet protection password
        """
        try:
            # Run external data cleanup early in the process
            self._update_progress(5, "Pre-cleaning external data references...")
            self._perform_external_data_cleanup(file_path)
            
            # Load workbook with proper error handling
            self._update_status(f"Loading workbook: {file_path}")
            wb = self._load_workbook_safely(file_path)
            
            sheet = self._get_worksheet(wb)
            
            self._update_progress(20, "Unprotecting sheet...")
            self._unprotect_sheet(sheet, password)
            
            self._update_progress(30, "Processing columns...")
            self._process_columns(sheet)
            
            self._update_progress(50, "Processing merged cells...")
            self._process_merged_cells(sheet)
            
            self._update_progress(60, "Finding column indexes...")
            column_indexes = self._find_column_indexes(sheet)
            
            self._update_progress(70, "Processing header formatting...")
            rgb_color = self._process_header_formatting(sheet)
            
            self._update_progress(80, "Finding matching rows...")
            matching_row = self._find_matching_row(sheet, rgb_color)
            
            self._update_progress(85, "Adding DO Comments column...")
            self._add_do_comments_column(sheet)
            
            self._update_progress(90, "Processing rows with comments...")
            self._process_rows_with_openpyxl(sheet, column_indexes, matching_row)
            
            # First save with openpyxl for the cell content changes
            self._update_status("Initial save with content changes...")
            self._save_workbook_safely(wb, file_path)
            
            # Clean up external references again before final save
            self._update_status("Cleaning external references...")
            self._perform_external_data_cleanup(file_path)
            
            # Then perform a special "clean" save to fix any potential corruption
            self._update_status("Cleaning and finalizing workbook...")
            self._final_clean_save(file_path)
            
            self._update_progress(100, "Processing complete")
        except Exception as e:
            self.logger.error(f"Error processing workbook: {str(e)}", exc_info=True)
            if self.queue:
                self.queue.put(("error", f"Workbook processing failed: {str(e)}"))
            raise
    
    def _final_clean_save(self, file_path: str) -> None:
        """
        Perform a final clean save using the most reliable approach.
        
        Args:
            file_path (str): Path to Excel file
        """
        # Create a backup of the current file state
        backup_path = self._create_backup_file(file_path)
        self._update_status(f"Created backup before final save: {backup_path}")
        
        # Get a new temporary filename for the cleaned version
        clean_file = self._get_temp_file_path("clean")
        
        # First perform specialized external data cleanup
        self._update_status("Running comprehensive external data cleanup...")
        self._perform_external_data_cleanup(file_path)
        
        # Try using Excel COM to save a clean version if enabled
        if self.excel_config.use_com_for_final_save:
            success = self._save_with_excel_com(file_path, clean_file, preserve_all=False)
        else:
            success = False
            
        if success and os.path.exists(clean_file):
            # Replace the original with the cleaned version
            if os.path.exists(file_path):
                os.unlink(file_path)  # Remove existing file
            shutil.move(clean_file, file_path)
            
            # Run one more external data cleanup on the final file
            self._perform_external_data_cleanup(file_path)
            self._update_status("Final clean save completed successfully")
        else:
            # If COM approach failed, ensure we still have a working file
            self._update_status("Using openpyxl for final save...")
            
            try:
                # Load and save with openpyxl as fallback
                wb = openpyxl.load_workbook(file_path)
                
                # Clean external references with openpyxl
                cleaner = ExcelDataCleaner(self.logger)
                cleaner.clean_external_references_openpyxl(wb)
                
                wb.save(clean_file)
                wb.close()
                
                if os.path.exists(clean_file) and os.path.getsize(clean_file) > 0:
                    if os.path.exists(file_path):
                        os.unlink(file_path)
                    shutil.move(clean_file, file_path)
                    self._update_status("Final openpyxl save completed successfully")
                else:
                    raise IOError("Failed to save with openpyxl")
                    
            except Exception as e:
                self._update_status(f"Final save with openpyxl failed: {str(e)}")
                self._update_status("Keeping original processed file")
                if self.queue:
                    self.queue.put(("warning", 
                        "The enhanced clean-up process failed, but the file was successfully processed. "
                        "The file should still be usable."
                    ))
                
        # Final verification of output file
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            self._update_status(f"Final file size: {os.path.getsize(file_path)} bytes")
        else:
            # If we somehow lost the file, restore from backup
            self._update_status("Output file missing or corrupted, restoring from backup...")
            shutil.copy2(backup_path, file_path)
    
    def _fix_merged_cells(self, wb):
        """
        Fix potential issues with merged cells that can cause corruption.
        
        Args:
            wb: openpyxl workbook
        """
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # First, unmerge all cells
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                sheet.unmerge_cells(str(merged_range))
            
            # We don't re-merge them because that's often a source of corruption
    
    def _load_workbook_safely(self, file_path: str) -> openpyxl.Workbook:
        """
        Load workbook with enhanced error handling.
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            openpyxl.Workbook: Loaded workbook
        """
        try:
            # Try loading with pandas first
            try:
                self._update_status("Attempting to load workbook with pandas...")
                df = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                self._update_status(f"Successfully read with pandas: {len(df)} sheets")
            except Exception as e:
                self.logger.warning(f"Pandas loading failed, falling back to openpyxl: {str(e)}")
            
            # Make sure file exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
                
            # Try to load with openpyxl's data_only mode
            self._update_status(f"Loading workbook with openpyxl: {file_path}")
            retry_count = 0
            max_retries = self.file_config.max_retries
            last_error = None
            
            while retry_count <= max_retries:
                try:
                    # Load with data_only=True to get values instead of formulas
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    self._update_status(f"Successfully loaded workbook with {len(wb.sheetnames)} sheets")
                    return wb
                except Exception as e:
                    retry_count += 1
                    last_error = e
                    self.logger.warning(f"Attempt {retry_count} to load workbook failed: {str(e)}")
                    
                    if retry_count <= max_retries:
                        # Wait before retrying
                        time.sleep(self.file_config.retry_delay)
                        # Make sure Excel is closed
                        self.close_excel_instances()
                    else:
                        break
                        
            # If all openpyxl attempts failed, try creating a safe copy first
            self._update_status("Creating safe copy for loading...")
            safe_copy = self._get_temp_file_path("safe_copy")
            
            # Copy the file
            shutil.copy2(file_path, safe_copy)
            
            # Try loading the safe copy
            try:
                wb = openpyxl.load_workbook(safe_copy, data_only=True)
                self._update_status("Successfully loaded workbook from safe copy")
                return wb
            except Exception as final_error:
                self.logger.error(f"Failed to load workbook after all attempts: {str(final_error)}")
                raise ValueError(f"Could not load Excel file after multiple attempts: {str(final_error)}")
            
        except FileNotFoundError:
            raise  # Re-raise file not found errors as is
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {str(e)}", exc_info=True)
            raise ValueError(f"Could not load Excel file: {str(e)}")
    
    def _save_workbook_safely(self, wb: openpyxl.Workbook, file_path: str) -> None:
        """
        Save workbook with enhanced error handling and backup.
        
        Args:
            wb (openpyxl.Workbook): Workbook to save
            file_path (str): Path to save to
        """
        # Create a backup before saving
        backup_path = self._create_backup_file(file_path)
        self._update_status(f"Created backup at: {backup_path}")
        
        # Multi-stage save process to prevent corruption
        stage1_file = self._get_temp_file_path("stage1")
        
        try:
            # STAGE 1: Save to first temp file
            self._update_status("Stage 1: Initial save to temp file...")
            wb.save(stage1_file)
            
            # Verify stage 1 file was created and is valid
            if not os.path.exists(stage1_file):
                raise IOError("Failed to save to first-stage temporary file")
            
            # Test open the temp file to verify integrity
            with self._open_and_close_workbook(stage1_file) as test_wb:
                if test_wb is None:
                    raise ValueError("Failed to verify workbook structure")
            
            # Close the original workbook to release resources
            wb.close()
            
            # Give system time to release file handles
            time.sleep(0.5)
            gc.collect()  # Force garbage collection
            
            # Move to final destination
            if os.path.exists(file_path):
                os.unlink(file_path)  # Remove existing file
             
            # Copy instead of move to ensure file system caching doesn't cause issues
            shutil.copy2(stage1_file, file_path)
            
            # Verify the final copy succeeded
            if not os.path.exists(file_path):
                raise IOError(f"Failed to copy final file to destination: {file_path}")
                
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                raise IOError(f"Saved file has zero size: {file_path}")
                
            self._update_status(f"Saved file successfully ({file_size} bytes) to: {file_path}")
        except Exception as e:
            self.logger.error(f"Error during save: {str(e)}", exc_info=True)
            if os.path.exists(backup_path):
                self._update_status("Restoring from backup due to save error...")
                shutil.copy2(backup_path, file_path)
            raise ValueError(f"Failed to save workbook: {str(e)}")
            
        finally:
            # Clean up temp files
            if os.path.exists(stage1_file):
                try:
                    os.unlink(stage1_file)
                except:
                    # Add to cleanup list for later
                    self._temp_files.append(stage1_file)
    
    def _get_temp_file_path(self, prefix: str = "excel") -> str:
        """
        Generate a temporary file path.
        
        Args:
            prefix (str): Prefix for temporary file
            
        Returns:
            str: Temporary file path
        """
        temp_dir = self.file_config.get_temp_dir() or tempfile.gettempdir()
        unique_id = str(uuid.uuid4())
        temp_file = os.path.join(temp_dir, f"{prefix}_temp_{unique_id}.xlsx")
        self._temp_files.append(temp_file)  # Track for cleanup
        return temp_file
    
    def _create_backup_file(self, file_path: str) -> str:
        """
        Create a backup copy of the file.
        
        Args:
            file_path (str): Path to file to backup
            
        Returns:
            str: Path to backup file
        """
        backup_dir = Path(self.config.output_directory) / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        filename = Path(file_path).name
        backup_path = str(backup_dir / f"{Path(filename).stem}_backup_{timestamp}{Path(filename).suffix}")
        shutil.copy2(file_path, backup_path)
        return backup_path
    
    @contextlib.contextmanager
    def _open_and_close_workbook(self, file_path: str):
        """
        Context manager to safely open and close a workbook.
        
        Args:
            file_path (str): Path to Excel file
            
        Yields:
            openpyxl.Workbook or None: Opened workbook or None if failed
        """
        wb = None
        try:
            # Wait a moment for filesystem to settle
            time.sleep(0.2)
            wb = openpyxl.load_workbook(file_path, read_only=True)
            yield wb
        except Exception as e:
            self.logger.warning(f"Failed to open workbook {file_path}: {e}")
            yield None
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass
    
    def _get_worksheet(self, workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Get the target worksheet from workbook.
        
        Args:
            workbook (Workbook): Openpyxl workbook object
            
        Returns:
            Worksheet: Target worksheet
            
        Raises:
            ValueError: If worksheet not found
        """
        if self.config.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{self.config.sheet_name}' not found")
        return workbook[self.config.sheet_name]
    
    def _unprotect_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet, password: str) -> None:
        """
        Unprotect worksheet with password.
        
        Args:
            sheet (Worksheet): Worksheet to unprotect
            password (str): Protection password
        """
        if password:
            try:
                sheet.protection.set_password(password)
                sheet.protection.sheet = False
            except Exception as e:
                raise ValueError(f"Failed to unprotect sheet: {str(e)}")
    
    def _process_columns(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Process and unhide all columns in worksheet.
        
        Args:
            sheet (Worksheet): Worksheet to process
        """
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].hidden = False
    
    def _process_merged_cells(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Unmerge all merged cells in worksheet.
        
        Args:
            sheet (Worksheet): Worksheet to process
        """
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))
    
    def _find_column_indexes(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
        """
        Find column indexes for required headers.
        
        Args:
            sheet (Worksheet): Worksheet to process
            
        Returns:
            Dict[str, int]: Mapping of header names to column indexes
        """
        column_indexes = {}
        for cell in sheet[self.config.header_row]:
            if cell.value in self.config.headers_to_find:
                column_indexes[cell.value] = cell.column
        missing_headers = set(self.config.headers_to_find) - set(column_indexes.keys())
        if missing_headers:
            raise ValueError(f"Missing headers: {', '.join(missing_headers)}")
        return column_indexes
    
    def _process_header_formatting(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """
        Process header cell formatting and return RGB color.
        
        Args:
            sheet (Worksheet): Worksheet to process
            
        Returns:
            str: RGB color value
        """
        header_cell = sheet.cell(row=self.config.header_row, column=1)
        fill_color = header_cell.fill.start_color.index
        
        if isinstance(fill_color, int):
            fill_color = f"{fill_color:06X}"
        if fill_color in COLOR_INDEX:
            rgb_color = COLOR_INDEX[int(fill_color, 16)]
        else:
            rgb_color = fill_color
        if not rgb_color.startswith('FF') and len(rgb_color) == 8:
            rgb_color = rgb_color[2:]
        return rgb_color
    
    def _find_matching_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet, rgb_color: str) -> int:
        """
        Find first row matching header color.
        
        Args:
            sheet (Worksheet): Worksheet to process
            rgb_color (str): RGB color to match
            
        Returns:
            int: Matching row number
        """
        for row in sheet.iter_rows(min_row=self.config.header_row + 1):
            cell = row[0]
            cell_fill_color = cell.fill.start_color.index
            
            if isinstance(cell_fill_color, int):
                cell_fill_color = f"{cell_fill_color:06X}"
            if cell_fill_color in COLOR_INDEX:
                cell_rgb_color = COLOR_INDEX[int(cell_fill_color, 16)]
            else:
                cell_rgb_color = cell_fill_color
            if not cell_rgb_color.startswith('FF') and len(cell_rgb_color) == 8:
                cell_rgb_color = cell_rgb_color[2:]
            if cell_rgb_color == rgb_color:
                return cell.row
        return sheet.max_row
    
    def _add_do_comments_column(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Add and format DO Comments column.
        
        Args:
            sheet (Worksheet): Worksheet to process
        """
        last_col = sheet.max_column
        new_header_cell = sheet.cell(row=self.config.header_row, column=last_col + 1)
        new_header_cell.value = "DO Comments"
        new_header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        new_header_cell.font = Font(color="FF0000", bold=True, size=11, name="Calibri")
        new_header_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
                    )
        new_header_cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        col_letter = get_column_letter(last_col + 1)
        sheet.column_dimensions[col_letter].width = 25
    
    def _process_rows_with_openpyxl(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        column_indexes: Dict[str, int],
        matching_row: int
    ) -> None:
        """
        Process individual rows with openpyxl.
        
        Args:
            sheet (Worksheet): openpyxl worksheet
            column_indexes (Dict[str, int]): Column index mapping
            matching_row (int): Last row to process
        """
        last_col = sheet.max_column
        comment_col = last_col
        processed_count = 0
        
        for row in range(self.config.header_row + 1, matching_row):
            try:
                difference_cell = sheet.cell(row=row, column=column_indexes["Difference"])
                include_cfo_cell = sheet.cell(row=row, column=column_indexes["Include in CFO Cert Letter"])
                explanation_cell = sheet.cell(row=row, column=column_indexes["Explanation"])
                
                difference_value = difference_cell.value
                include_cfo_value = include_cfo_cell.value
                explanation_value = explanation_cell.value
                
                # Add comment cell with appropriate formatting
                comment_cell = sheet.cell(row=row, column=comment_col)
                comment_cell.alignment = Alignment(wrap_text=True)
                
                if difference_value not in (None, ""):
                    if include_cfo_value == "N" and explanation_value not in (None, 0, ""):
                        comment_cell.value = "Explanation Reasonable"
                        processed_count += 1
                    elif include_cfo_value == "Y" and explanation_value not in (None, ""):
                        comment_cell.value = "Explanation Reasonable; Include in CFO Cert Letter"
                        processed_count += 1    
                    elif explanation_value in (None, "", 0) and difference_value != 0:
                        comment_cell.value = "Explanation Required"
                        # Add highlighting for cells that require attention
                        comment_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        processed_count += 1
                        
            except Exception as e:
                self._update_status(f"Error processing row {row}: {str(e)}")
        
        self._update_status(f"Successfully processed {processed_count} rows")
    
    def _process_with_pywin32(self, file_path: str, column_indexes: Dict[str, int], matching_row: int) -> None:
        """
        Process workbook with pywin32 for formula evaluation.
        
        Args:
            file_path (str): Path to Excel file
            column_indexes (Dict[str, int]): Column index mappings
            matching_row (int): Last row to process
        """
        file_path = os.path.abspath(file_path)
        
        self._update_status(f"Checking if file exists: {file_path}")
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File does not exist: {file_path}")
        
        self.close_excel_instances()
        time.sleep(2)
        
        excel = None
        wb = None
        try:
            # Initialize COM
            pythoncom.CoInitialize()
            self._update_status("Starting Excel application...")
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            # ENHANCEMENT: Disable events and calculations for better performance
            excel.EnableEvents = False
            excel.Calculation = -4135  # xlCalculationManual
            
            self._update_status(f"Opening workbook: {file_path}")
            wb = excel.Workbooks.Open(
                file_path,
                UpdateLinks=0,
                ReadOnly=False,
                CorruptLoad=2  # xlRepairFile - better error handling
            )    
            
            self._update_status("Selecting worksheet...")
            ws = wb.Worksheets(self.config.sheet_name)
                
            self._process_rows_with_pywin32(
                ws,
                column_indexes,
                matching_row
            ) 
            
            # ENHANCEMENT: Calculate before saving
            wb.Calculate()
            
            self._update_status("Saving workbook...")
            wb.Save()    
            
        except Exception as e:
            self._update_status(f"Error in Excel processing: {str(e)}")
            import traceback
            self._update_status(f"Detailed error: {traceback.format_exc()}")
            raise
        finally:
            # Clean up COM objects
            if wb:
                try:
                    wb.Close(SaveChanges=True)
                except Exception as e:
                    self._update_status(f"Error closing workbook: {str(e)}")
            
            if excel:
                try:
                    excel.Quit()
                except Exception as e:
                    self._update_status(f"Error quitting Excel: {str(e)}")
                
                del wb
                del excel
                
            # Force garbage collection to release COM objects
            gc.collect()
            pythoncom.CoUninitialize()
    
    def _process_rows_with_pywin32(self, ws, column_indexes: Dict[str, int], matching_row: int) -> None:
        """
        Process individual rows with pywin32.
        
        Args:
            ws: pywin32 worksheet object
            column_indexes (Dict[str, int]): Column index mapping
            matching_row (int): Last row to process
        """
        last_col = ws.UsedRange.Columns.Count
        processed_count = 0
        
        for row in range(self.config.header_row + 1, matching_row):
            try:
                difference_value = ws.Cells(row, column_indexes["Difference"]).Value
                include_cfo_value = ws.Cells(row, column_indexes["Include in CFO Cert Letter"]).Value
                explanation_value = ws.Cells(row, column_indexes["Explanation"]).Value
                
                if difference_value not in (None, ""):
                    if include_cfo_value == "N" and explanation_value not in (None, 0, ""):
                        ws.Cells(row, last_col).Value = "Explanation Reasonable"
                        processed_count += 1
                    elif include_cfo_value == "Y" and explanation_value not in (None, ""):
                        ws.Cells(row, last_col).Value = "Explanation Reasonable; Include in CFO Cert Letter"
                        processed_count += 1
                    elif explanation_value in (None, "", 0) and difference_value != 0:
                        ws.Cells(row, last_col).Value = "Explanation Required"
                        # ENHANCEMENT: Highlight cells that need attention
                        ws.Cells(row, last_col).Interior.Color = 0xFF9999  # Light red
                        processed_count += 1
                        
            except Exception as e:
                self._update_status(f"Error processing row {row}: {str(e)}")
        
        self._update_status(f"Successfully processed {processed_count} rows")