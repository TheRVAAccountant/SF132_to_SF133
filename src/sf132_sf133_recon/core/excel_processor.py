"""
Core Excel processor for SF132/SF133 reconciliation.

This module contains the primary ExcelProcessor class that handles the main
Excel processing operations for SF132 to SF133 reconciliation.
"""

import os
import time
import logging
import gc
import shutil
import tempfile
import uuid
from pathlib import Path
from typing import Dict, Tuple, Optional, Any, List, Union
from queue import Queue

# Import from the modules package to avoid circular imports
try:
    from ..modules.excel_handler import close_excel_instances as handler_close_excel_instances
except ImportError:
    handler_close_excel_instances = None

# Try to import Excel-related modules - they might be unavailable on some platforms
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.cell import MergedCell
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import InvalidFileException
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import Windows-specific modules - these are essential for the application
import sys

# Check if running on Windows
IS_WINDOWS = sys.platform.startswith('win')

# Import Windows-specific modules
if IS_WINDOWS:
    try:
        import pythoncom
        import win32com.client
        import win32api
        WINDOWS_COM_AVAILABLE = True
    except ImportError as e:
        # On Windows, these imports are critical
        raise ImportError(f"Critical Windows modules missing: {e}. Please install 'pywin32' package.") from e
else:
    # When not on Windows, note that functionality will be limited
    WINDOWS_COM_AVAILABLE = False
    print("WARNING: Not running on Windows - Excel COM automation unavailable. Functionality will be limited.")

# Local configuration
DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_HEADER_ROW = 5
DEFAULT_HEADERS_TO_FIND = ["Difference", "Include in CFO Cert Letter", "Explanation"]
DEFAULT_OUTPUT_DIR = "output"

class ExcelProcessor:
    """
    Handles Excel file processing operations including file manipulation,
    formatting, and content analysis for SF132 to SF133 reconciliation.
    """
    
    def __init__(self, queue: Optional[Queue] = None):
        """
        Initialize the Excel processor.
        
        Args:
            queue: Queue for communication with GUI
        """
        self.queue = queue
        self._setup_logging()
        self._temp_files = []  # Track temp files for cleanup
        
        # Default configuration
        self.sheet_name = DEFAULT_SHEET_NAME
        self.header_row = DEFAULT_HEADER_ROW
        self.headers_to_find = DEFAULT_HEADERS_TO_FIND
        self.output_directory = DEFAULT_OUTPUT_DIR
        
        # Create output directory
        os.makedirs(self.output_directory, exist_ok=True)
        
    def _setup_logging(self):
        """Configure logging for the processor."""
        self.logger = logging.getLogger(__name__)
    
    def __del__(self):
        """Clean up resources when instance is destroyed."""
        self._cleanup_temp_files()
        
    def _cleanup_temp_files(self):
        """Clean up all temporary files."""
        for temp_file in self._temp_files:
            if os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                    self.logger.debug(f"Cleaned up temp file: {temp_file}")
                except Exception as e:
                    self.logger.warning(f"Failed to clean up temp file {temp_file}: {e}")
    
    def _update_progress(self, value: float, message: str):
        """
        Update progress through the queue.
        
        Args:
            value: Progress value (0-100)
            message: Status message
        """
        if self.queue:
            self.queue.put(("progress", (value, message)))
        self.logger.info(message)
    
    def _update_status(self, message: str):
        """
        Send status update through the queue.
        
        Args:
            message: Status message
        """
        if self.queue:
            self.queue.put(("status", message))
        self.logger.info(message)
    
    def close_excel_instances(self):
        """Terminate all existing Excel processes to prevent file locking."""
        self._update_status("Ensuring all Excel instances are closed...")
        
        # Use the module version if available
        if handler_close_excel_instances:
            try:
                handler_close_excel_instances()
                self._update_status("Excel instances closed via handler module")
                return
            except Exception as e:
                self.logger.warning(f"Handler close_excel_instances failed: {e}")
                # Fall through to local implementation
        
        try:
            import psutil
            excel_pids = []
            
            # First find all Excel processes, including COM components
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    # Check both process name and command line for Excel references
                    proc_name = proc.info.get('name', '').upper()
                    cmd_line = ' '.join(proc.info.get('cmdline', [])).upper() if proc.info.get('cmdline') else ''
                    
                    # Look for any Excel-related processes
                    if any(excel_marker in proc_name for excel_marker in ['EXCEL', 'EXCEL.EXE', 'MICROSOFT EXCEL']) or \
                       any(excel_marker in cmd_line for excel_marker in ['EXCEL.EXE', 'EXCELCNV', 'DCOM']):
                        excel_pids.append(proc.info['pid'])
                except Exception:
                    pass
            
            if excel_pids:
                self._update_status(f"Found {len(excel_pids)} Excel-related processes to close")
                
            # First try graceful termination with COM cleanup
            if IS_WINDOWS and WINDOWS_COM_AVAILABLE:
                try:
                    # Initialize COM
                    pythoncom.CoInitialize()
                    
                    # Try to quit any remaining Excel applications through COM
                    try:
                        excel = win32com.client.GetActiveObject("Excel.Application")
                        excel.DisplayAlerts = False
                        excel.Quit()
                        del excel
                        gc.collect()
                        self.logger.info("Gracefully closed active Excel application via COM")
                    except:
                        pass
                    
                    # Uninitialize COM
                    pythoncom.CoUninitialize()
                except:
                    pass
                    
            # Then terminate each process with proper cleanup
            for pid in excel_pids:
                try:
                    proc = psutil.Process(pid)
                    proc.terminate()
                    proc.wait(timeout=5)
                except Exception as e:
                    self.logger.warning(f"Failed to terminate Excel process (PID {pid}): {e}")
                    # Try more aggressive termination
                    try:
                        if psutil.pid_exists(pid):
                            os.kill(pid, 9)  # SIGKILL or equivalent on Windows
                    except Exception:
                        pass
            
            # Verify all processes are terminated
            remaining = []
            for pid in excel_pids:
                if psutil.pid_exists(pid):
                    remaining.append(pid)
                    
            if remaining:
                self.logger.warning(f"Could not terminate {len(remaining)} Excel processes: {remaining}")
                # For stubborn processes, try the Windows-specific taskkill command
                if IS_WINDOWS:
                    try:
                        import subprocess
                        self._update_status("Using system taskkill command for stubborn Excel processes...")
                        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                                      stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                    except Exception as e:
                        self.logger.warning(f"Taskkill failed: {e}")
            else:
                self._update_status("All Excel processes successfully closed")
                
            # Add a delay to ensure file system has time to release locks
            time.sleep(2)
            
            # Force garbage collection to release COM objects
            gc.collect()
            
        except ImportError:
            self.logger.warning("psutil module not available, cannot close Excel instances")

    def process_file(self, original_file: str, password: str = None) -> bool:
        """
        Main processing function for Excel file.
        
        Args:
            original_file: Path to original Excel file
            password: Sheet protection password
        
        Returns:
            bool: True if processing was successful, False otherwise
        """
        try:
            # Convert to absolute path
            original_file = os.path.abspath(original_file)
            self._update_status(f"Processing file: {original_file}")
            
            # Maximum number of attempts for the overall process
            max_attempts = 3
            attempt_count = 0
            
            # Create a verified backup before any processing
            try:
                backup_dir = Path(self.output_directory) / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                
                # Create a timestamped backup
                timestamp = time.strftime("%Y%m%d-%H%M%S")
                backup_name = f"{Path(original_file).stem}_original_backup_{timestamp}.xlsx"
                backup_path = str(backup_dir / backup_name)
                
                # Copy the original file
                shutil.copy2(original_file, backup_path)
                self._update_status(f"Created original backup at: {backup_path}")
            except Exception as e:
                self.logger.warning(f"Could not create original backup: {e}")
                # Continue even if backup fails - this is just an extra safety measure
            
            while attempt_count < max_attempts:
                try:
                    attempt_count += 1
                    self._update_status(f"Processing attempt {attempt_count}/{max_attempts}")
                    
                    # Basic validation of input file
                    self._validate_file(original_file)
                    new_file = self._generate_new_filename(original_file)
                    
                    # Ensure Excel is fully closed before starting
                    self.close_excel_instances()
                    
                    # First make a clean copy of the original file with additional error handling
                    try:
                        temp_copy = self._create_verified_copy(original_file)
                    except Exception as copy_error:
                        self.logger.warning(f"Verified copy failed: {copy_error}, using direct copy")
                        # Fall back to direct copy if verified copy fails
                        temp_copy = self._get_temp_file_path("direct_temp_copy")
                        
                        # Use low-level file copy for reliability
                        try:
                            # Use a larger buffer for faster copying
                            buffer_size = 10 * 1024 * 1024  # 10MB buffer
                            
                            with open(original_file, 'rb') as src, open(temp_copy, 'wb') as dst:
                                while True:
                                    chunk = src.read(buffer_size)
                                    if not chunk:
                                        break
                                    dst.write(chunk)
                        except Exception:
                            # Fall back to simple copy
                            shutil.copy2(original_file, temp_copy)
                        
                        # Verify the copy exists and has content
                        if not os.path.exists(temp_copy) or os.path.getsize(temp_copy) == 0:
                            raise ValueError("Failed to create a valid copy of the original file")
                    
                    # Main processing approaches in order of preference:
                    # 1. Fresh workbook approach - most reliable but can fail with complex files
                    # 2. Legacy library-based processing - more compatible but slower
                    # 3. Legacy general processing - most compatible with complex files
                    
                    # APPROACH 1: Try the fresh workbook approach first (most reliable)
                    self._update_progress(10, "Creating fresh workbook...")
                    success = self._process_with_fresh_workbook(temp_copy, new_file, password)
                    
                    if success:
                        self._update_progress(95, "Validating final workbook...")
                        if self._validate_excel_file(new_file):
                            self._update_progress(100, "Processing complete")
                            self._update_status(f"Successfully created and processed: {new_file}")
                            if self.queue:
                                self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                            return True
                        else:
                            # If validation fails, try repair process
                            self._update_status("File validation failed, attempting repair...")
                            if self._repair_workbook(new_file):
                                self._update_status("File repaired successfully")
                                if self.queue:
                                    self.queue.put(("success", f"File processed and repaired. Output saved to: {new_file}"))
                                return True
                            else:
                                # If repair failed but this is the last attempt, continue with the file anyway
                                if attempt_count >= max_attempts:
                                    self._update_status("Repair failed but providing file anyway")
                                    if self.queue:
                                        self.queue.put(("warning", f"File processed with warnings. Output saved to: {new_file}"))
                                    return True
                                
                                # Otherwise, try next approach
                                self.logger.warning("Repair failed, trying next approach")
                                raise ValueError("Failed to create a valid Excel file")
                    
                    # APPROACH 2: Fall back to legacy processing methods
                    self._update_status("Fresh workbook approach failed, falling back to library methods...")
                    success = self._process_with_legacy_method(original_file, new_file, password)
                    
                    if success:
                        self._update_progress(100, "Legacy processing complete")
                        self._update_status(f"Successfully processed with legacy method: {new_file}")
                        if self.queue:
                            self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                        return True
                        
                    # If everything failed, try again with a delay
                    self.logger.warning(f"All processing attempts failed in round {attempt_count}")
                    
                    if attempt_count < max_attempts:
                        # Clear resources before next attempt
                        self.close_excel_instances()
                        gc.collect()
                        time.sleep(3)  # Wait before next attempt
                        
                        # Special error handling for subsequent attempts
                        if attempt_count == 2:
                            # On the last attempt, try more aggressive resource cleanup
                            self._update_status("Performing aggressive resource cleanup before final attempt...")
                            
                            # Force garbage collection multiple times
                            for _ in range(3):
                                gc.collect()
                                time.sleep(1)
                            
                            # On Windows, try to find and kill any hidden Excel processes
                            if IS_WINDOWS:
                                try:
                                    import subprocess
                                    # Use taskkill with force option
                                    subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE', '/T'], 
                                                  stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                                except:
                                    pass
                    
                except Exception as attempt_error:
                    self.logger.error(f"Attempt {attempt_count} failed: {attempt_error}", exc_info=True)
                    
                    if attempt_count >= max_attempts:
                        # If we've exhausted all attempts, report the error
                        if self.queue:
                            self.queue.put(("error", f"Failed after {max_attempts} attempts: {str(attempt_error)}"))
                        return False
                    
                    # Clean up and prepare for next attempt
                    self.close_excel_instances()
                    gc.collect()
                    time.sleep(2)
            
            # If we get here, all attempts have failed
            if self.queue:
                self.queue.put(("error", f"Failed after {max_attempts} attempts"))
            return False
                
        except Exception as e:
            self.logger.error("Processing failed", exc_info=True)
            if self.queue:
                self.queue.put(("error", str(e)))
            return False
        finally:
            # Final cleanup
            self._cleanup_temp_files()
    
    def _create_verified_copy(self, original_file: str) -> str:
        """
        Create a verified copy of the original file to prevent corruption.
        
        Args:
            original_file: Path to original Excel file
            
        Returns:
            str: Path to verified copy
        """
        temp_copy = self._get_temp_file_path("verified_copy")
        self._update_status(f"Creating verified copy at {temp_copy}...")
        
        # Use native Excel to create a clean copy (most reliable method)
        try:
            if WINDOWS_COM_AVAILABLE:
                # Start Excel
                pythoncom.CoInitialize()
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                
                # Open workbook with recovery options
                wb = excel.Workbooks.Open(
                    original_file,
                    UpdateLinks=0,
                    ReadOnly=True,
                    CorruptLoad=2  # xlRepairFile (better corruption handling)
                )
                
                # Save as a new clean file
                wb.SaveAs(
                    temp_copy,
                    FileFormat=51,  # xlOpenXMLWorkbook
                    CreateBackup=False
                )
                
                # Clean close
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                # Force cleanup
                del wb
                del excel
                gc.collect()
                pythoncom.CoUninitialize()
                
                # Verify file exists and has content
                if not os.path.exists(temp_copy) or os.path.getsize(temp_copy) == 0:
                    raise ValueError("Failed to create valid copy")
                    
                return temp_copy
            else:
                # If COM is not available, use regular copy
                shutil.copy2(original_file, temp_copy)
                return temp_copy
                
        except Exception as e:
            self.logger.warning(f"COM copy failed: {e}, falling back to direct copy")
            
            # Fallback to direct copy
            shutil.copy2(original_file, temp_copy)
            return temp_copy
    
    def _validate_excel_file(self, file_path: str) -> bool:
        """
        Validate an Excel file by trying to open it and check for errors.
        
        Args:
            file_path: Path to Excel file to validate
            
        Returns:
            bool: Whether the file is valid
        """
        try:
            # Method 1: Try opening with openpyxl
            if OPENPYXL_AVAILABLE:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                self._update_status(f"Basic validation with openpyxl successful. Found {len(sheet_names)} sheets.")
            
            # Method 2: Verify with Excel COM if possible
            if WINDOWS_COM_AVAILABLE:
                # Make sure Excel is not running
                self.close_excel_instances()
                
                # Initialize COM with error handling
                try:
                    pythoncom.CoInitialize()
                except Exception as e:
                    self.logger.warning(f"COM initialization failed: {e}")
                
                excel = None
                wb = None
                
                try:
                    # Start Excel with proper exception handling
                    try:
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                    except Exception as e:
                        self.logger.warning(f"Excel COM initialization failed: {e}")
                        return True  # Fall back to openpyxl validation
                    
                    # Try to open the file
                    try:
                        wb = excel.Workbooks.Open(
                            file_path,
                            UpdateLinks=0,
                            ReadOnly=True
                        )
                    except Exception as e:
                        self.logger.warning(f"Excel COM file open failed: {e}")
                        return True  # Fall back to openpyxl validation
                    
                    # Check for errors using a more reliable method than ErrorCheckingStatus
                    has_errors = False
                    
                    # Try different error checking methods
                    try:
                        # Method 1: Check if ErrorCheckingStatus attribute exists and use it
                        if hasattr(excel, 'ErrorCheckingStatus'):
                            has_errors = excel.ErrorCheckingStatus
                        # Method 2: Use Excel's CheckWorkbookCompatibility method if available
                        elif hasattr(wb, 'CheckCompatibility'):
                            compatibility_issues = wb.CheckCompatibility
                            has_errors = compatibility_issues
                        # Method 3: Check for error alerts on sheets
                        else:
                            # If no direct error checking is available, check if file can be saved
                            temp_path = self._get_temp_file_path("verify")
                            wb.SaveAs(temp_path, FileFormat=51)  # 51 = xlsx
                            if os.path.exists(temp_path):
                                os.unlink(temp_path)
                            has_errors = False
                    except Exception as e:
                        self.logger.debug(f"Error checking method failed: {e}")
                        # If error checking fails, assume the file is okay if we got this far
                        has_errors = False
                    
                    # Close properly
                    if wb:
                        try:
                            wb.Close(SaveChanges=False)
                        except:
                            pass
                        
                    if excel:
                        try:
                            excel.Quit()
                        except:
                            pass
                        
                    # Cleanup COM objects
                    if wb:
                        del wb
                    if excel:
                        del excel
                        
                    gc.collect()
                    
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                    
                    if has_errors:
                        self.logger.warning(f"Excel detected errors in {file_path}")
                        return False
                    
                    self._update_status("Enhanced COM validation successful")
                    
                except Exception as e:
                    self.logger.warning(f"COM validation failed: {e}")
                    # Clean up resources on error
                    if wb:
                        try:
                            wb.Close(SaveChanges=False)
                        except:
                            pass
                    if excel:
                        try:
                            excel.Quit()
                        except:
                            pass
                        
                    # Force cleanup
                    if wb:
                        del wb
                    if excel:
                        del excel
                    gc.collect()
                    
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                    
                    return True  # Fall back to openpyxl validation
            
            # Additional validation with pandas if available
            if PANDAS_AVAILABLE:
                try:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_count = len(excel_file.sheet_names)
                    excel_file.close()
                    self._update_status(f"Validated file with pandas: {sheet_count} sheets")
                except Exception as e:
                    self.logger.warning(f"Pandas validation failed: {e}")
            
            # If we got here, file seems valid
            return True
            
        except Exception as e:
            self.logger.warning(f"File validation failed: {e}")
            return False
    
    def _process_with_fresh_workbook(self, source_file: str, output_file: str, password: str) -> bool:
        """
        Process by extracting data and creating a fresh workbook to avoid corruption.
        
        Args:
            source_file: Source Excel file
            output_file: Path for output file
            password: Sheet protection password
            
        Returns:
            bool: Whether processing succeeded
        """
        try:
            self._update_status("Loading source workbook data...")
            
            # Load source workbook
            if not OPENPYXL_AVAILABLE:
                return False
                
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            
            if self.sheet_name not in source_wb.sheetnames:
                self.logger.warning(f"Required sheet '{self.sheet_name}' not found")
                return False
                
            source_sheet = source_wb[self.sheet_name]
            
            # Create a new workbook
            self._update_status("Creating fresh workbook...")
            new_wb = openpyxl.Workbook()
            
            # Remove default sheet
            if "Sheet" in new_wb.sheetnames:
                new_wb.remove(new_wb["Sheet"])
                
            # Create new sheet with the right name
            new_sheet = new_wb.create_sheet(title=self.sheet_name)
            
            # Copy all data and formatting from source
            self._update_status("Copying data from source...")
            self._copy_sheet_data(source_sheet, new_sheet)
            
            # Process the fresh sheet
            self._update_progress(40, "Processing data...")
            
            # Process column visibility
            self._process_columns(new_sheet)
            
            # Find column indexes
            column_indexes = self._find_column_indexes(new_sheet)
            
            # Find header color
            rgb_color = self._process_header_formatting(new_sheet)
            
            # Find matching row
            matching_row = self._find_matching_row(new_sheet, rgb_color)
            
            # Add DO Comments column
            self._add_do_comments_column(new_sheet)
            
            # Process rows with comments
            self._process_rows_with_openpyxl(new_sheet, column_indexes, matching_row)
            
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Save the workbook
            self._update_status("Saving processed workbook...")
            new_wb.save(output_file)
            
            # Cleanup
            source_wb.close()
            new_wb.close()
            
            # Verify the saved file
            if not os.path.exists(output_file):
                return False
                
            return True
            
        except Exception as e:
            self.logger.error(f"Error in fresh workbook processing: {e}", exc_info=True)
            return False
    
    def _copy_sheet_data(self, source_sheet, target_sheet):
        """
        Copy data and basic formatting from source sheet to target sheet.
        
        Args:
            source_sheet: Source worksheet
            target_sheet: Target worksheet
        """
        # Get dimensions of source sheet
        max_row = source_sheet.max_row
        max_col = source_sheet.max_column
        
        # Copy column dimensions and properties
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_sheet.column_dimensions:
                src_col_dim = source_sheet.column_dimensions[col_letter]
                tgt_col_dim = target_sheet.column_dimensions[col_letter]
                
                # Copy column width
                tgt_col_dim.width = src_col_dim.width if src_col_dim.width else 8.43  # Default width
                
                # Copy hidden status
                tgt_col_dim.hidden = False  # We want all columns visible
        
        # Copy row heights
        for row_idx in range(1, max_row + 1):
            if row_idx in source_sheet.row_dimensions:
                src_row_dim = source_sheet.row_dimensions[row_idx]
                tgt_row_dim = target_sheet.row_dimensions[row_idx]
                
                # Copy row height
                tgt_row_dim.height = src_row_dim.height if src_row_dim.height else 15  # Default height
        
        # Copy cell data and basic formatting
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                # Get source cell
                src_cell = source_sheet.cell(row=row_idx, column=col_idx)
                
                # Skip merged cells (we'll handle them separately)
                if isinstance(src_cell, MergedCell):
                    continue
                    
                # Create target cell and copy value
                tgt_cell = target_sheet.cell(row=row_idx, column=col_idx, value=src_cell.value)
                
                # Copy basic formatting with proper error handling
                if hasattr(src_cell, 'has_style') and src_cell.has_style:
                    try:
                        # Font
                        tgt_cell.font = Font(
                            name=src_cell.font.name,
                            size=src_cell.font.size,
                            bold=src_cell.font.bold,
                            italic=src_cell.font.italic,
                            color=src_cell.font.color
                        )
                    except Exception as e:
                        self.logger.debug(f"Error copying font at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Fill - with proper handling to avoid black fills
                        if src_cell.fill and hasattr(src_cell.fill, 'start_color') and src_cell.fill.start_color:
                            # Get the fill type from source
                            fill_type = getattr(src_cell.fill, 'fill_type', 'solid')
                            
                            # Skip empty fills
                            if fill_type == 'none' or fill_type is None:
                                continue
                                
                            # Get the RGB color
                            fill_color = src_cell.fill.start_color.rgb
                            
                            # Skip black colors (000000) or None values
                            if not fill_color or fill_color == "00000000" or fill_color == "FF000000" or fill_color == "000000":
                                continue
                                
                            # Default to white if color is invalid
                            if not isinstance(fill_color, str) or len(fill_color) < 6:
                                fill_color = "FFFFFF"
                                
                            # Ensure proper format by removing alpha channel if present
                            if len(fill_color) == 8 and fill_color.startswith("FF"):
                                fill_color = fill_color[2:]
                                
                            # Apply the fill directly with string color (more reliable than Color object)
                            tgt_cell.fill = PatternFill(
                                fill_type='solid',
                                start_color=fill_color
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying fill at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Border
                        if src_cell.border:
                            tgt_cell.border = Border(
                                left=src_cell.border.left,
                                right=src_cell.border.right,
                                top=src_cell.border.top,
                                bottom=src_cell.border.bottom
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying border at {row_idx},{col_idx}: {e}")
                    
                    try:
                        # Alignment
                        if src_cell.alignment:
                            tgt_cell.alignment = Alignment(
                                horizontal=src_cell.alignment.horizontal,
                                vertical=src_cell.alignment.vertical,
                                wrap_text=src_cell.alignment.wrap_text
                            )
                    except Exception as e:
                        self.logger.debug(f"Error copying alignment at {row_idx},{col_idx}: {e}")
    
    def _repair_workbook(self, file_path: str) -> bool:
        """
        Repair a potentially corrupted Excel file.
        
        Args:
            file_path: Path to Excel file to repair
            
        Returns:
            bool: Whether repair was successful
        """
        try:
            self._update_status("Attempting to repair workbook...")
            
            # Ensure all Excel processes are closed
            self.close_excel_instances()
            
            # Create a backup before repair
            backup_path = None
            try:
                backup_dir = Path(self.output_directory) / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                
                # Create a timestamped backup
                timestamp = time.strftime("%Y%m%d-%H%M%S")
                backup_name = f"{Path(file_path).stem}_repair_backup_{timestamp}.xlsx"
                backup_path = str(backup_dir / backup_name)
                
                # Copy the file
                shutil.copy2(file_path, backup_path)
                self._update_status(f"Created backup at: {backup_path}")
            except Exception as e:
                self.logger.warning(f"Could not create backup: {e}")
            
            # Check file access before attempting repair
            max_retries = 5
            retry_count = 0
            file_accessible = False
            
            while retry_count < max_retries and not file_accessible:
                try:
                    # Test file accessibility with read/write checks
                    with open(file_path, 'rb+') as test_file:
                        test_file.seek(0)
                        test_file.read(1)  # Try to read a byte
                        test_file.seek(0)  # Go back to beginning
                    file_accessible = True
                except Exception as e:
                    retry_count += 1
                    self.logger.warning(f"File access retry {retry_count}/{max_retries}: {e}")
                    
                    # Try to release the file by forcing garbage collection
                    gc.collect()
                    time.sleep(2)  # Wait before retrying
                    
                    # On Windows, try additional methods to release the file
                    if IS_WINDOWS:
                        self.close_excel_instances()
                        
                        # Try to work with a copy if original is inaccessible
                        if retry_count >= 3:
                            try:
                                temp_copy = self._get_temp_file_path("repair_copy")
                                self._update_status(f"Attempting to create a copy at {temp_copy}...")
                                
                                # Use low-level file operations to copy
                                with open(file_path, 'rb') as src, open(temp_copy, 'wb') as dst:
                                    dst.write(src.read())
                                
                                # If copy succeeds, work with the copy instead
                                if os.path.exists(temp_copy) and os.path.getsize(temp_copy) > 0:
                                    self._update_status("Working with copy since original is locked")
                                    file_path = temp_copy
                                    file_accessible = True
                            except Exception as copy_err:
                                self.logger.warning(f"Failed to create copy: {copy_err}")
            
            if not file_accessible:
                self.logger.error("Could not access file for repair after multiple attempts")
                return False
            
            # Try with module recovery first if available
            try:
                from ..modules.excel_recovery import repair_excel_file_access
                self._update_status("Using module-based recovery...")
                success, result_path = repair_excel_file_access(file_path)
                if success and os.path.exists(result_path) and os.path.getsize(result_path) > 0:
                    # Validate the result
                    if self._validate_repaired_file(result_path):
                        # Replace original with repaired version
                        if os.path.exists(file_path) and os.path.abspath(file_path) != os.path.abspath(result_path):
                            os.unlink(file_path)
                        shutil.copy2(result_path, file_path)
                        self._update_status("Module-based repair successful")
                        return True
            except ImportError:
                self.logger.warning("Excel recovery module not available")
            except Exception as e:
                self.logger.warning(f"Module-based repair failed: {e}")
            
            # Try multiple repair methods in sequence
            repair_methods = [
                self._repair_with_excel_com,
                self._repair_with_pandas,
                self._repair_with_openpyxl,
                self._repair_with_system_tool
            ]
            
            # Try each repair method until one succeeds
            for repair_method in repair_methods:
                try:
                    self._update_status(f"Trying repair method: {repair_method.__name__}")
                    success, repaired_path = repair_method(file_path)
                    
                    if success and os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                        # Validate the repaired file
                        if self._validate_repaired_file(repaired_path):
                            # Replace original with repaired version
                            try:
                                # Enhanced file replacement logic
                                original_filename = os.path.basename(file_path)
                                self._update_status(f"Replacing {original_filename} with repaired version...")
                                
                                # Remove original if it exists and isn't the same as the repaired file
                                if os.path.exists(file_path) and os.path.abspath(file_path) != os.path.abspath(repaired_path):
                                    try:
                                        os.unlink(file_path)
                                    except Exception as e:
                                        self.logger.warning(f"Could not remove original file: {e}")
                                        # If we can't delete, try to move it instead
                                        try:
                                            failed_path = self._get_temp_file_path("failed_original")
                                            shutil.move(file_path, failed_path)
                                        except:
                                            pass
                                
                                # Copy the repaired file to the original location
                                shutil.copy2(repaired_path, file_path)
                                
                                # Verify the final file
                                if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                                    self._update_status("Repair completed successfully")
                                    return True
                            except Exception as e:
                                self.logger.warning(f"File replacement failed: {e}")
                                # Continue to next method if replacement fails
                        else:
                            self.logger.warning("Repaired file validation failed")
                except Exception as method_error:
                    self.logger.warning(f"{repair_method.__name__} failed: {method_error}")
            
            # If all repair methods failed, try to restore from backup
            if backup_path and os.path.exists(backup_path) and os.path.getsize(backup_path) > 0:
                self._update_status("All repair methods failed. Attempting to restore from backup...")
                try:
                    # Ensure target file is accessible
                    if os.path.exists(file_path):
                        try:
                            os.unlink(file_path)
                        except:
                            # If deletion fails, try to use a new path instead
                            file_path = self._get_temp_file_path("restored")
                    
                    # Copy the backup
                    shutil.copy2(backup_path, file_path)
                    
                    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                        self._update_status("Successfully restored from backup")
                        return True
                except Exception as restore_error:
                    self.logger.error(f"Backup restoration failed: {restore_error}")
            
            return False
            
        except Exception as e:
            self.logger.error(f"Repair process failed: {e}")
            return False
            
    def _validate_repaired_file(self, file_path: str) -> bool:
        """
        Validate a repaired Excel file to ensure it's usable.
        
        Args:
            file_path: Path to the repaired file
            
        Returns:
            bool: Whether the file is valid
        """
        try:
            # Check if file exists and has size
            if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                return False
                
            # Try opening with openpyxl in read-only mode
            if OPENPYXL_AVAILABLE:
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    sheet_count = len(wb.sheetnames)
                    wb.close()
                    self._update_status(f"Validated repaired file with {sheet_count} sheets")
                    return True
                except Exception as e:
                    self.logger.warning(f"openpyxl validation failed: {e}")
                
            # Try with pandas as fallback
            if PANDAS_AVAILABLE:
                try:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_count = len(excel_file.sheet_names)
                    excel_file.close()
                    self._update_status(f"Validated repaired file with pandas: {sheet_count} sheets")
                    return True
                except Exception as e:
                    self.logger.warning(f"pandas validation failed: {e}")
                
            return False
            
        except Exception as e:
            self.logger.warning(f"Repair validation failed: {e}")
            return False
            
    def _repair_with_excel_com(self, file_path: str) -> tuple:
        """
        Repair Excel file using Excel's COM interface.
        
        Args:
            file_path: Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not IS_WINDOWS or not WINDOWS_COM_AVAILABLE:
            return False, None
            
        excel = None
        wb = None
        repaired_path = self._get_temp_file_path("com_repaired")
        
        try:
            # Ensure Excel is closed
            self.close_excel_instances()
            
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Start Excel with robust error handling
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.EnableEvents = False
                excel.Calculation = -4135  # xlCalculationManual
            except Exception as e:
                self.logger.warning(f"Excel COM initialization failed: {e}")
                return False, None
                
            # Try to open the file in repair mode
            try:
                self._update_status("Opening file in repair mode...")
                wb = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=0,
                    CorruptLoad=2  # xlRepairFile - repair mode
                )
            except Exception as e:
                self.logger.warning(f"Failed to open file in repair mode: {e}")
                # Try with different parameters
                try:
                    wb = excel.Workbooks.Open(
                        file_path,
                        UpdateLinks=0,
                        ReadOnly=True
                    )
                except Exception as e2:
                    self.logger.warning(f"Failed to open file in read-only mode: {e2}")
                    return False, None
            
            # Calculate workbook
            try:
                wb.Calculate()
            except:
                pass
                
            # Save to a new file
            self._update_status(f"Saving repaired file to {repaired_path}...")
            wb.SaveAs(
                repaired_path,
                FileFormat=51,  # xlOpenXMLWorkbook
                CreateBackup=False
            )
            
            # Close workbook and quit Excel
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Cleanup
            del wb
            del excel
            wb = None
            excel = None
            gc.collect()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Excel COM repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Excel COM repair failed: {e}")
            return False, None
        finally:
            # Ensure proper cleanup
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
                    
            # Force garbage collection
            if wb:
                del wb
            if excel:
                del excel
            gc.collect()
            
            # Uninitialize COM
            try:
                pythoncom.CoUninitialize()
            except:
                pass
                
    def _repair_with_pandas(self, file_path: str) -> tuple:
        """
        Repair Excel file using pandas.
        
        Args:
            file_path: Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not PANDAS_AVAILABLE:
            return False, None
            
        repaired_path = self._get_temp_file_path("pandas_repaired")
        
        try:
            self._update_status("Attempting pandas-based repair...")
            
            # Try to read the file with pandas
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # Create a new Excel writer
            writer = pd.ExcelWriter(repaired_path, engine='openpyxl')
            
            # Process each sheet
            for sheet_name in sheet_names:
                self._update_status(f"Processing sheet: {sheet_name}")
                # Read with error handling for individual sheets
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as sheet_err:
                    self.logger.warning(f"Could not process sheet {sheet_name}: {sheet_err}")
                    # Create an empty sheet instead
                    df = pd.DataFrame()
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Save the writer
            writer.close()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Pandas repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Pandas repair failed: {e}")
            return False, None
            
    def _repair_with_openpyxl(self, file_path: str) -> tuple:
        """
        Repair Excel file using openpyxl.
        
        Args:
            file_path: Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not OPENPYXL_AVAILABLE:
            return False, None
            
        repaired_path = self._get_temp_file_path("openpyxl_repaired")
        
        try:
            self._update_status("Attempting openpyxl-based repair...")
            
            # Try to read the workbook
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
            except Exception as read_err:
                self.logger.warning(f"Failed to open with openpyxl: {read_err}")
                return False, None
                
            # Create a new workbook
            new_wb = openpyxl.Workbook()
            
            # Remove default sheet
            if "Sheet" in new_wb.sheetnames:
                new_wb.remove(new_wb["Sheet"])
                
            # Copy each sheet from source
            for sheet_name in wb.sheetnames:
                self._update_status(f"Processing sheet: {sheet_name}")
                # Create new sheet
                new_sheet = new_wb.create_sheet(title=sheet_name)
                src_sheet = wb[sheet_name]
                
                # Try to copy basic data
                try:
                    # Get rows from source sheet
                    for row_idx, row in enumerate(src_sheet.iter_rows(values_only=True), 1):
                        for col_idx, value in enumerate(row, 1):
                            new_sheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception as sheet_err:
                    self.logger.warning(f"Error copying sheet {sheet_name}: {sheet_err}")
            
            # Close the source workbook
            wb.close()
            
            # Save the new workbook
            new_wb.save(repaired_path)
            new_wb.close()
            
            # Verify the repaired file
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                self._update_status("Openpyxl repair successful")
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"Openpyxl repair failed: {e}")
            return False, None
            
    def _repair_with_system_tool(self, file_path: str) -> tuple:
        """
        Try to repair the Excel file using system tools (Windows only).
        
        Args:
            file_path: Path to Excel file to repair
            
        Returns:
            tuple: (success, repaired_file_path)
        """
        if not IS_WINDOWS:
            return False, None
            
        repaired_path = self._get_temp_file_path("system_repaired")
        
        try:
            self._update_status("Attempting system-based repair...")
            
            # First try to copy the file to a new location
            shutil.copy2(file_path, repaired_path)
            
            # Then try ExcelCnv command-line tool if available (on some Windows systems)
            try:
                import subprocess
                office_paths = [
                    r"C:\Program Files\Microsoft Office\root\Office16",
                    r"C:\Program Files (x86)\Microsoft Office\root\Office16",
                    r"C:\Program Files\Microsoft Office\Office16",
                    r"C:\Program Files (x86)\Microsoft Office\Office16",
                    r"C:\Program Files\Microsoft Office\Office15",
                    r"C:\Program Files (x86)\Microsoft Office\Office15",
                ]
                
                excelcnv_path = None
                for office_path in office_paths:
                    test_path = os.path.join(office_path, "excelcnv.exe")
                    if os.path.exists(test_path):
                        excelcnv_path = test_path
                        break
                        
                if excelcnv_path:
                    self._update_status(f"Found Excel converter at: {excelcnv_path}")
                    
                    # Generate a temporary output path
                    output_path = self._get_temp_file_path("excelcnv_output")
                    
                    # Run ExcelCnv to convert/repair the file
                    subprocess.run([
                        excelcnv_path,
                        "-nme",  # No message boxes
                        "-oice",  # Open Invalid with Converter Extensions
                        "-xlsb",  # Convert to XLSB format first (more robust)
                        repaired_path,
                        output_path
                    ], stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                    
                    # Convert back to xlsx
                    if os.path.exists(output_path):
                        final_path = self._get_temp_file_path("final_repaired")
                        subprocess.run([
                            excelcnv_path,
                            "-nme",
                            "-xlsx",  # Convert to XLSX format
                            output_path,
                            final_path
                        ], stderr=subprocess.PIPE, stdout=subprocess.PIPE)
                        
                        if os.path.exists(final_path) and os.path.getsize(final_path) > 0:
                            self._update_status("System repair successful")
                            return True, final_path
            except Exception as system_err:
                self.logger.warning(f"System tool repair failed: {system_err}")
            
            # If system tools failed, return the simple copy if it exists
            if os.path.exists(repaired_path) and os.path.getsize(repaired_path) > 0:
                return True, repaired_path
                
            return False, None
            
        except Exception as e:
            self.logger.warning(f"System repair failed: {e}")
            return False, None
    
    def _process_with_legacy_method(self, original_file: str, new_file: str, password: str = None) -> bool:
        """
        Process using the legacy method as a fallback.
        
        Args:
            original_file: Original file path
            new_file: New file path
            password: Sheet protection password
            
        Returns:
            bool: Whether processing was successful
        """
        try:
            # First try to process with pandas/openpyxl directly
            self._update_progress(5, "Attempting to process file with pandas/openpyxl...")
            try:
                success = self._process_with_libraries(original_file, new_file, password)
                if success:
                    self._update_progress(90, "Direct library processing successful")
                    self._update_status(f"Successfully created and processed: {new_file}")
                    if self.queue:
                        self.queue.put(("success", f"File processed successfully. Output saved to: {new_file}"))
                    return True
                else:
                    self._update_status("Direct library processing failed, falling back to basic copy...")
            except Exception as e:
                self.logger.warning(f"Direct library processing failed: {e}")
                self._update_status("Falling back to basic file copy...")
            
            # Very simple fallback - just make a copy
            try:
                self._update_progress(10, f"Creating new workbook at {new_file}...")
                shutil.copy2(original_file, new_file)
                
                self._update_progress(100, "Processing complete (basic copy only)")
                self._update_status(f"Successfully created copy: {new_file}")
                if self.queue:
                    self.queue.put(("success", f"File copied successfully. Output saved to: {new_file}"))
                return True
            except Exception as e:
                self.logger.error(f"Basic copy failed: {str(e)}", exc_info=True)
                if self.queue:
                    self.queue.put(("error", f"Processing failed: {str(e)}"))
                return False
                
        except Exception as e:
            self.logger.error("Processing failed", exc_info=True)
            if self.queue:
                self.queue.put(("error", str(e)))
            return False
    
    def _direct_file_copy(self, source_file: str, dest_file: str) -> bool:
        """
        Perform a direct file copy without using COM.
        This is more reliable when COM operations fail.
        
        Args:
            source_file: Source file path
            dest_file: Destination file path
            
        Returns:
            bool: Whether the copy was successful
        """
        try:
            self._update_status(f"Directly copying file from {source_file} to {dest_file}...")
            
            # Ensure the destination directory exists
            os.makedirs(os.path.dirname(dest_file), exist_ok=True)
            
            # Make sure we can access the source file
            max_retries = 5
            retry_count = 0
            file_accessible = False
            
            while retry_count < max_retries and not file_accessible:
                try:
                    # Test if we can read the source file
                    with open(source_file, 'rb') as test_file:
                        test_file.read(1)  # Try to read a byte
                    file_accessible = True
                except Exception as e:
                    retry_count += 1
                    self.logger.warning(f"Source file access retry {retry_count}/{max_retries}: {e}")
                    
                    # Try to release the file by forcing garbage collection
                    gc.collect()
                    time.sleep(2)  # Wait before retrying
                    
                    # Make sure Excel is closed
                    if retry_count >= 2:
                        self.close_excel_instances()
            
            if not file_accessible:
                self.logger.error("Could not access source file for copying")
                return False
            
            # Check for destination file accessibility
            if os.path.exists(dest_file):
                try:
                    # Try to remove existing dest file if it might be locked
                    with open(dest_file, 'r+b') as test_file:
                        test_file.seek(0)
                except Exception as e:
                    self.logger.warning(f"Destination file is locked, trying to use a different path: {e}")
                    # Use a new destination path
                    old_dest = dest_file
                    dest_file = self._get_temp_file_path("direct_copy")
                    self._update_status(f"Using alternative path: {dest_file}")
            
            # Try several copy methods
            copy_methods = [
                # Method 1: Standard shutil.copy2
                lambda s, d: shutil.copy2(s, d),
                
                # Method 2: Low-level file copy with chunks
                lambda s, d: self._chunk_copy(s, d),
                
                # Method 3: Use system commands (Windows)
                lambda s, d: self._system_copy(s, d) if IS_WINDOWS else None,
                
                # Method 4: os.system copy (fallback)
                lambda s, d: os.system(f'copy "{s}" "{d}"') if IS_WINDOWS else None
            ]
            
            # Try each copy method until one succeeds
            success = False
            for method_idx, copy_method in enumerate(copy_methods):
                try:
                    self._update_status(f"Trying copy method {method_idx+1}...")
                    result = copy_method(source_file, dest_file)
                    if result is not None and result != 0:
                        continue  # Try next method if this one returned an error code
                    
                    # Verify the copy succeeded
                    if os.path.exists(dest_file) and os.path.getsize(dest_file) > 0:
                        file_size = os.path.getsize(dest_file)
                        self._update_status(f"Direct file copy successful ({file_size} bytes)")
                        success = True
                        break
                except Exception as e:
                    self.logger.warning(f"Copy method {method_idx+1} failed: {e}")
                    continue
            
            if not success:
                self.logger.error("All copy methods failed")
                return False
                
            return True
            
        except Exception as e:
            self.logger.error(f"Error during direct file copy: {str(e)}", exc_info=True)
            return False
            
    def _chunk_copy(self, source_file: str, dest_file: str) -> bool:
        """
        Copy a file in chunks to avoid memory issues with large files.
        
        Args:
            source_file: Source file path
            dest_file: Destination file path
            
        Returns:
            bool: Whether the copy was successful
        """
        try:
            # Use a larger buffer for faster copying of large files
            buffer_size = 10 * 1024 * 1024  # 10MB buffer
            
            with open(source_file, 'rb') as src, open(dest_file, 'wb') as dst:
                while True:
                    chunk = src.read(buffer_size)
                    if not chunk:
                        break
                    dst.write(chunk)
            
            return True
        except Exception as e:
            self.logger.warning(f"Chunk copy failed: {e}")
            return False
    
    def _system_copy(self, source_file: str, dest_file: str) -> int:
        """
        Use system-specific copy commands for reliability.
        
        Args:
            source_file: Source file path
            dest_file: Destination file path
            
        Returns:
            int: Return code (0 for success)
        """
        if IS_WINDOWS:
            # Windows - use robocopy or xcopy
            try:
                import subprocess
                # Try robocopy first (more reliable for locked files)
                source_dir = os.path.dirname(source_file)
                source_file_name = os.path.basename(source_file)
                dest_dir = os.path.dirname(dest_file)
                
                # Use /R:3 to retry 3 times if file is locked, /W:2 to wait 2 seconds between retries
                # /J for unbuffered I/O, /B for backup mode (can copy open files), /NP for no progress
                result = subprocess.run(
                    ['robocopy', source_dir, dest_dir, source_file_name, '/R:3', '/W:2', '/J', '/B', '/NP'],
                    capture_output=True,
                    text=True
                )
                
                # Robocopy return codes: 0 = no files copied, 1 = files copied, > 1 = errors
                if result.returncode <= 1:
                    return 0  # Success
                
                # If robocopy failed, try xcopy
                result = subprocess.run(
                    ['xcopy', source_file, dest_file, '/Y', '/Q', '/R', '/H'],
                    capture_output=True,
                    text=True
                )
                
                return result.returncode
            except Exception as e:
                self.logger.warning(f"System copy failed: {e}")
                return 1
        else:
            # Unix-like - use cp
            try:
                import subprocess
                result = subprocess.run(['cp', source_file, dest_file], capture_output=True, text=True)
                return result.returncode
            except Exception as e:
                self.logger.warning(f"System cp failed: {e}")
                return 1
    
    def _create_clean_copy(self, original_file: str, new_file: str) -> bool:
        """
        Create a clean copy of the original file using COM if possible.
        
        Args:
            original_file: Original file path
            new_file: New file path
            
        Returns:
            bool: Whether the operation was successful
        """
        try:
            # First make sure the output directory exists
            os.makedirs(os.path.dirname(new_file), exist_ok=True)
            
            # Create a temporary file for intermediate processing
            temp_file = self._get_temp_file_path("init_copy")
            
            # Try opening the original file to verify its integrity
            self._update_status("Verifying original file integrity...")
            
            # Use COM if available
            if WINDOWS_COM_AVAILABLE and IS_WINDOWS:
                try:
                    # Initialize COM
                    pythoncom.CoInitialize()
                    
                    # Start Excel
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    
                    # Open workbook with recovery options
                    wb = excel.Workbooks.Open(
                        original_file,
                        UpdateLinks=0,
                        ReadOnly=True,
                        CorruptLoad=2  # xlRepairFile (better corruption handling)
                    )
                    
                    # Save as a new clean file
                    wb.SaveAs(
                        temp_file,
                        FileFormat=51,  # xlOpenXMLWorkbook
                        CreateBackup=False
                    )
                    
                    # Clean close
                    wb.Close(SaveChanges=False)
                    excel.Quit()
                    
                    # Force cleanup
                    del wb
                    del excel
                    gc.collect()
                    pythoncom.CoUninitialize()
                    
                    # Verify file exists and has content
                    if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
                        raise ValueError("Failed to create valid copy")
                    
                    # Copy the temp file to the final destination
                    shutil.copy2(temp_file, new_file)
                    
                    # Clean external connections
                    try:
                        from ..modules.excel_cleaner import clean_excel_external_data
                        clean_excel_external_data(new_file, self.logger)
                    except ImportError:
                        # Try inline cleaning if module is not available
                        if WINDOWS_COM_AVAILABLE and IS_WINDOWS:
                            self._clean_external_connections(new_file)
                    
                    return True
                    
                except Exception as e:
                    self.logger.warning(f"COM copy failed: {e}, falling back to direct copy")
            
            # Direct copy as fallback
            if self._direct_file_copy(original_file, new_file):
                self._update_status("Created copy successfully using direct file copy")
                return True
            else:
                self.logger.error("Failed to create copy")
                return False
                
        except Exception as e:
            self.logger.error(f"Error during clean copy creation: {str(e)}")
            return False
    
    def _clean_external_connections(self, file_path: str) -> bool:
        """
        Clean external connections in the Excel file to prevent security issues.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            bool: Whether the cleaning was successful
        """
        if not WINDOWS_COM_AVAILABLE or not IS_WINDOWS:
            return False
            
        excel = None
        wb = None
        
        try:
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Start Excel with limited automation security
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open workbook
            wb = excel.Workbooks.Open(
                file_path,
                UpdateLinks=0,
                ReadOnly=False
            )
            
            # Break all external links
            try:
                # Excel constants for link types
                xlExcelLinks = 1
                xlOLELinks = 2
                xlPublishers = 5
                xlTextImport = 6
                
                for link_type in [xlExcelLinks, xlOLELinks, xlPublishers, xlTextImport]:
                    try:
                        wb.BreakLink(Name="", Type=link_type)
                    except:
                        pass
            except:
                pass
            
            # Remove any data connections
            if hasattr(wb, 'Connections'):
                try:
                    for i in range(wb.Connections.Count, 0, -1):
                        try:
                            wb.Connections(i).Delete()
                        except:
                            pass
                except:
                    pass
            
            # Save the workbook
            wb.Save()
            
            # Close properly
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            return True
            
        except Exception as e:
            self.logger.warning(f"External connection cleanup failed: {e}")
            return False
        finally:
            # Cleanup
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass
            
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
            
            # Force cleanup of COM objects
            if wb:
                del wb
            if excel:
                del excel
            gc.collect()
            
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    
    def _process_with_libraries(self, original_file: str, output_file: str, password: str) -> bool:
        """
        Process Excel file using Python libraries (openpyxl/pandas) instead of COM.
        
        Args:
            original_file: Path to original Excel file
            output_file: Path to output file
            password: Sheet protection password
            
        Returns:
            bool: True if processing was successful, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl not available, cannot process with libraries")
            return False
            
        try:
            self._update_status(f"Loading file with pandas/openpyxl: {original_file}")
            
            # Create output directory if needed
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Copy the file first for safety
            self._direct_file_copy(original_file, output_file)
            
            # Load workbook with openpyxl
            self._update_status("Loading workbook with openpyxl...")
            wb = openpyxl.load_workbook(output_file, data_only=True)
            
            if self.sheet_name not in wb.sheetnames:
                self.logger.warning(f"Required sheet '{self.sheet_name}' not found")
                return False
            
            sheet = wb[self.sheet_name]
            
            # Unprotect sheet if password provided
            if password:
                self._update_progress(20, "Unprotecting sheet...")
                self._unprotect_sheet(sheet, password)
            
            self._update_progress(30, "Processing columns...")
            self._process_columns(sheet)
            
            self._update_progress(60, "Finding column indexes...")
            column_indexes = self._find_column_indexes(sheet)
            
            self._update_progress(70, "Processing header formatting...")
            rgb_color = self._process_header_formatting(sheet)
            
            self._update_progress(80, "Finding matching rows...")
            matching_row = self._find_matching_row(sheet, rgb_color)
            
            self._update_progress(85, "Adding DO Comments column...")
            self._add_do_comments_column(sheet)
            
            self._update_progress(90, "Processing rows with comments...")
            self._process_rows_with_openpyxl(sheet, column_indexes, matching_row)
            
            # Save the workbook
            wb.save(output_file)
            wb.close()
            
            self._update_progress(100, "Processing complete")
            return True
                
        except Exception as e:
            self.logger.warning(f"Library-based processing failed: {str(e)}")
            return False
            
    def _process_columns(self, sheet) -> None:
        """
        Process and unhide all columns in worksheet.
        
        Args:
            sheet: Worksheet to process
        """
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].hidden = False
    
    def _unprotect_sheet(self, sheet, password: str) -> None:
        """
        Unprotect worksheet with password.
        
        Args:
            sheet: Worksheet to unprotect
            password: Protection password
        """
        if password:
            try:
                if hasattr(sheet.protection, 'set_password'):
                    sheet.protection.set_password(password)
                sheet.protection.sheet = False
            except Exception as e:
                self.logger.warning(f"Failed to unprotect sheet: {e}")
    
    def _find_column_indexes(self, sheet) -> Dict[str, int]:
        """
        Find column indexes for required headers.
        
        Args:
            sheet: Worksheet to process
            
        Returns:
            Dict[str, int]: Mapping of header names to column indexes
        """
        column_indexes = {}
        
        # Set up default headers if not already specified
        headers_to_find = getattr(self, 'headers_to_find', DEFAULT_HEADERS_TO_FIND)
        header_row = getattr(self, 'header_row', DEFAULT_HEADER_ROW)
        
        # Find all headers
        for cell in sheet[header_row]:
            if cell.value in headers_to_find:
                column_indexes[cell.value] = cell.column
                
        # Check if we found all required headers
        if len(column_indexes) < len(headers_to_find):
            self.logger.warning(f"Not all required headers found. Found: {list(column_indexes.keys())}")
            
        return column_indexes
    
    def _process_header_formatting(self, sheet) -> str:
        """
        Process header cell formatting and return RGB color.
        
        Args:
            sheet: Worksheet to process
            
        Returns:
            str: RGB color value
        """
        header_row = getattr(self, 'header_row', DEFAULT_HEADER_ROW)
        header_cell = sheet.cell(row=header_row, column=1)
        
        # Default to white if no fill
        if not hasattr(header_cell, 'fill') or not header_cell.fill:
            return "FFFFFF"
            
        # Extract fill color
        fill_color = getattr(header_cell.fill.start_color, 'index', "FFFFFF")
        
        # Convert index to RGB if needed
        if isinstance(fill_color, int):
            fill_color = f"{fill_color:06X}"
            
        # Handle color index
        rgb_color = fill_color
        if hasattr(COLOR_INDEX, '__contains__') and int(fill_color, 16) in COLOR_INDEX:
            rgb_color = COLOR_INDEX[int(fill_color, 16)]
            
        # Clean up color format
        if isinstance(rgb_color, str) and len(rgb_color) == 8 and rgb_color.startswith('FF'):
            rgb_color = rgb_color[2:]
            
        return rgb_color
    
    def _find_matching_row(self, sheet, rgb_color: str) -> int:
        """
        Find first row matching header color.
        
        Args:
            sheet: Worksheet to process
            rgb_color: RGB color to match
            
        Returns:
            int: Matching row number
        """
        header_row = getattr(self, 'header_row', DEFAULT_HEADER_ROW)
        
        for row in sheet.iter_rows(min_row=header_row + 1):
            cell = row[0]
            
            # Skip cells without fill
            if not hasattr(cell, 'fill') or not cell.fill:
                continue
                
            # Get cell fill color
            cell_fill_color = getattr(cell.fill.start_color, 'index', None)
            
            # Skip cells without fill color
            if not cell_fill_color:
                continue
                
            # Convert to RGB
            if isinstance(cell_fill_color, int):
                cell_fill_color = f"{cell_fill_color:06X}"
                
            # Handle color index
            cell_rgb_color = cell_fill_color
            if hasattr(COLOR_INDEX, '__contains__') and int(cell_fill_color, 16) in COLOR_INDEX:
                cell_rgb_color = COLOR_INDEX[int(cell_fill_color, 16)]
                
            # Clean up color format
            if isinstance(cell_rgb_color, str) and len(cell_rgb_color) == 8 and cell_rgb_color.startswith('FF'):
                cell_rgb_color = cell_rgb_color[2:]
                
            # Compare colors
            if cell_rgb_color == rgb_color:
                return cell.row
                
        # If no matching row found, return max row
        return sheet.max_row
    
    def _add_do_comments_column(self, sheet) -> None:
        """
        Add and format DO Comments column.
        
        Args:
            sheet: Worksheet to process
        """
        last_col = sheet.max_column
        header_row = getattr(self, 'header_row', DEFAULT_HEADER_ROW)
        
        # Create the header cell
        new_header_cell = sheet.cell(row=header_row, column=last_col + 1)
        new_header_cell.value = "DO Comments"
        
        # Format the header cell with explicit colors (no Color objects)
        new_header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        new_header_cell.font = Font(color="FF0000", bold=True, size=11, name="Calibri")
        new_header_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        new_header_cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # Set column width
        col_letter = get_column_letter(last_col + 1)
        sheet.column_dimensions[col_letter].width = 25
        
        # Log the successful addition of DO Comments column
        self.logger.info(f"Added DO Comments column at position {last_col + 1} (column {col_letter})")
    
    def _process_rows_with_openpyxl(self, sheet, column_indexes: Dict[str, int], matching_row: int) -> None:
        """
        Process individual rows with openpyxl.
        
        Args:
            sheet: openpyxl worksheet
            column_indexes: Column index mapping
            matching_row: Last row to process
        """
        # Get the last column - this should be the DO Comments column we added
        last_col = sheet.max_column
        comment_col = last_col
        processed_count = 0
        header_row = getattr(self, 'header_row', DEFAULT_HEADER_ROW)
        
        # Verify DO Comments column exists, if not add it now
        header_cell = sheet.cell(row=header_row, column=comment_col)
        if header_cell.value != "DO Comments":
            self.logger.warning("DO Comments column not found where expected, attempting to add it now")
            self._add_do_comments_column(sheet)
            # Update the comment column pointer to the newly added column
            comment_col = sheet.max_column
            
        # Check if we have required columns
        required_columns = ["Difference", "Include in CFO Cert Letter", "Explanation"]
        missing_columns = [col for col in required_columns if col not in column_indexes]
        
        if missing_columns:
            self.logger.warning(f"Missing required columns: {missing_columns}")
            return
            
        self.logger.info(f"Processing rows from {header_row + 1} to {matching_row} with DO Comments in column {comment_col}")
        
        # Track the number of comments in each category
        explanation_reasonable_count = 0
        include_in_cfo_count = 0
        explanation_required_count = 0
            
        for row in range(header_row + 1, matching_row):
            try:
                # Get the cells for required columns
                difference_cell = sheet.cell(row=row, column=column_indexes["Difference"])
                include_cfo_cell = sheet.cell(row=row, column=column_indexes["Include in CFO Cert Letter"])
                explanation_cell = sheet.cell(row=row, column=column_indexes["Explanation"])
                
                # Extract values, ensuring we normalize to appropriate types
                try:
                    difference_value = float(difference_cell.value) if difference_cell.value not in (None, "") else 0
                except (ValueError, TypeError):
                    difference_value = 0
                    
                include_cfo_value = str(include_cfo_cell.value).upper() if include_cfo_cell.value is not None else ""
                explanation_value = explanation_cell.value
                
                # Prepare the comment cell
                comment_cell = sheet.cell(row=row, column=comment_col)
                
                # Default formatting for all comment cells
                comment_cell.alignment = Alignment(wrap_text=True, vertical='center')
                comment_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Clear any existing fills or values
                comment_cell.value = None
                comment_cell.fill = PatternFill(fill_type=None)
                
                # Apply business rules to determine comments
                if difference_value != 0:  # Only process rows with differences
                    if include_cfo_value in ("N", "NO") and explanation_value not in (None, "", 0):
                        comment_cell.value = "Explanation Reasonable"
                        explanation_reasonable_count += 1
                        processed_count += 1
                    elif include_cfo_value in ("Y", "YES") and explanation_value not in (None, "", 0):
                        comment_cell.value = "Explanation Reasonable; Include in CFO Cert Letter"
                        include_in_cfo_count += 1
                        processed_count += 1    
                    elif explanation_value in (None, "", 0):
                        comment_cell.value = "Explanation Required"
                        # Add highlighting for cells that require attention - explicit color values
                        comment_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        explanation_required_count += 1
                        processed_count += 1
                        
            except Exception as e:
                self.logger.warning(f"Error processing row {row}: {str(e)}")
        
        self.logger.info(f"DO Comments summary: {explanation_reasonable_count} Explanation Reasonable, "
                         f"{include_in_cfo_count} Include in CFO, {explanation_required_count} Explanation Required")
        self._update_status(f"Successfully processed {processed_count} rows with DO Comments")
    
    def _validate_file(self, file_path: str) -> None:
        """
        Validate the input file path.
        
        Args:
            file_path: Path to Excel file
        
        Raises:
            ValueError: If file validation fails
        """
        if not os.path.exists(file_path):
            raise ValueError(f"File does not exist: {file_path}")
        if not file_path.lower().endswith('.xlsx'):
            raise ValueError("File must be an Excel (.xlsx) file")
    
    def _generate_new_filename(self, original_file: str) -> str:
        """
        Generate the new filename based on the original.
        
        Args:
            original_file: Original file path
        
        Returns:
            str: New file path
        """
        # Get original file basename without extension
        original_basename = Path(original_file).stem
        
        # Create output directory if it doesn't exist
        output_dir = Path(self.output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create a descriptive filename with timestamp
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        new_filename = f"{original_basename}_processed_{timestamp}.xlsx"
        
        # Construct full path
        new_path = output_dir / new_filename
        
        self._update_status(f"Generated output filename: {new_path}")
        return str(new_path.absolute())
    
    def _get_temp_file_path(self, prefix: str = "excel") -> str:
        """
        Generate a temporary file path.
        
        Args:
            prefix: Prefix for temporary file
            
        Returns:
            str: Temporary file path
        """
        # Try to get temp directory from configuration
        try:
            from ..modules.file_operations import get_temp_file_path
            temp_file = get_temp_file_path(prefix)
            self._temp_files.append(temp_file)  # Track for cleanup
            return temp_file
        except ImportError:
            # Fallback to local implementation if module is not available
            temp_dir = tempfile.gettempdir()
            unique_id = str(uuid.uuid4())
            temp_file = os.path.join(temp_dir, f"{prefix}_temp_{unique_id}.xlsx")
            self._temp_files.append(temp_file)  # Track for cleanup
            return temp_file