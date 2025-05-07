"""
Excel file recovery module.

This module provides functions to recover from Excel file access and corruption issues.
It handles Windows-specific "file in use" errors and offers repair strategies.
"""

import os
import sys
import time
import logging
import shutil
import tempfile
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any, Union, Callable

# Import platform-specific modules if available
try:
    import pythoncom
    import win32com.client
    WINDOWS_COM_AVAILABLE = True
except ImportError:
    WINDOWS_COM_AVAILABLE = False

# Initialize logger
logger = logging.getLogger(__name__)

# Type aliases
PathLike = Union[str, Path]

def is_windows_platform() -> bool:
    """
    Check if running on Windows platform.
    
    Returns:
        bool: True if running on Windows
    """
    return sys.platform.startswith('win')

def repair_workbook(file_path: str) -> bool:
    """
    Repair a potentially corrupted Excel file.
    
    Args:
        file_path: Path to Excel file to repair
        
    Returns:
        bool: Whether repair was successful
    """
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return False
        
    logger.info(f"Attempting to repair workbook: {file_path}")
    
    # Create a backup before attempting repair
    from .file_operations import create_backup_file
    backup_path = create_backup_file(file_path)
    
    # Method 1: Try with pandas
    try:
        import pandas as pd
        logger.info("Attempting pandas-based repair...")
        
        # Read all sheets with pandas
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        
        # Create a new Excel writer
        temp_path = f"{file_path}.temp.xlsx"
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        
        # Copy each sheet
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                logger.warning(f"Error repairing sheet {sheet_name}: {e}")
        
        # Save the writer
        writer.close()
        
        # Replace original with repaired version
        if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
            if os.path.exists(file_path):
                os.unlink(file_path)
            shutil.copy2(temp_path, file_path)
            os.unlink(temp_path)
            logger.info("Successfully repaired with pandas")
            return True
            
    except Exception as e:
        logger.warning(f"Pandas repair failed: {e}")
    
    # Method 2: Try with COM (Windows only)
    if WINDOWS_COM_AVAILABLE and is_windows_platform():
        try:
            logger.info("Attempting COM-based repair...")
            
            # Method 2: Use Excel's repair functionality
            from .excel_handler import close_excel_instances
            close_excel_instances()
            
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Start Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            
            # Set calculation to manual to avoid recalculation issues
            excel.Calculation = -4135  # xlCalculationManual
            
            # Open with repair option
            wb = excel.Workbooks.Open(
                file_path,
                UpdateLinks=0,
                CorruptLoad=2  # xlRepairFile - repair mode
            )
            
            # Calculate once to ensure data is current
            wb.Calculate()
            
            # Save to a new temp file
            temp_path = f"{file_path}.repaired.xlsx"
            wb.SaveAs(
                temp_path,
                FileFormat=51,  # xlOpenXMLWorkbook
                CreateBackup=False
            )
            
            # Close workbook and Excel
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Cleanup COM objects
            del wb
            del excel
            import gc
            gc.collect()
            pythoncom.CoUninitialize()
            
            # Replace original with repaired version
            if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
                if os.path.exists(file_path):
                    os.unlink(file_path)
                shutil.copy2(temp_path, file_path)
                os.unlink(temp_path)
                logger.info("Successfully repaired with COM")
                return True
                
        except Exception as e:
            logger.warning(f"COM repair failed: {e}")
    
    # Method 3: Try direct openpyxl repair as last resort
    try:
        import openpyxl
        logger.info("Attempting openpyxl repair...")
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Remove any external links
        if hasattr(wb, '_external_links'):
            wb._external_links = []
        
        # Fix any merged cell issues
        for sheet in wb.worksheets:
            # Get all merged cell ranges
            merged_ranges = list(sheet.merged_cells.ranges)
            # Unmerge all cells (common source of corruption)
            for merged_range in merged_ranges:
                sheet.unmerge_cells(str(merged_range))
        
        # Save to a new temp file
        temp_path = f"{file_path}.fixed.xlsx"
        wb.save(temp_path)
        wb.close()
        
        # Replace original with fixed version
        if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
            if os.path.exists(file_path):
                os.unlink(file_path)
            shutil.copy2(temp_path, file_path)
            os.unlink(temp_path)
            logger.info("Successfully repaired with openpyxl")
            return True
            
    except Exception as e:
        logger.warning(f"openpyxl repair failed: {e}")
    
    # If all repair methods failed, restore from backup (already done in excel_processor.py)
    logger.error("All repair methods failed")
    return False

def fix_file_in_use_error(file_path: str) -> Tuple[bool, str]:
    """
    Fix the "file in use" error for Excel files with advanced Windows-specific recovery.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple[bool, str]: (Success status, Result message or path)
    """
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"
    
    logger.info(f"Attempting to fix file-in-use error for: {file_path}")
    
    # Create a timestamp for the fixed file
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    
    # Create a copy path for the fixed file
    fixed_path = str(Path(file_path).with_stem(f"{Path(file_path).stem}_fixed_{timestamp}"))
    
    # Windows-specific path normalization
    if is_windows_platform():
        try:
            from ..utils.win_path_handler import normalize_windows_path
            file_path = normalize_windows_path(file_path)
            fixed_path = normalize_windows_path(fixed_path)
        except ImportError:
            logger.warning("Windows path handler not available")
    
    # Approach 1: Close Excel instances and try simple copy
    from .excel_handler import close_excel_instances
    close_excel_instances()
    
    # On Windows, try Windows-specific unlocking
    if is_windows_platform():
        try:
            from ..utils.win_api import force_file_unlock, cleanup_excel_temp_files
            
            # Clean up Excel temp files that might be causing locks
            cleanup_excel_temp_files()
            
            # Force unlock the file
            force_file_unlock(file_path)
        except ImportError:
            logger.warning("Windows API module not available")
    
    # Wait a moment for the OS to release the file
    time.sleep(1)
    
    # Approach 2: Try to create a copy with robust file operations
    try:
        from .file_operations import safe_file_copy
        if safe_file_copy(file_path, fixed_path, retries=5):
            logger.info(f"Successfully created fixed copy: {fixed_path}")
            return True, fixed_path
    except Exception as e:
        logger.warning(f"Standard file copy failed: {e}")
    
    # Approach 3: On Windows, try advanced recovery
    if is_windows_platform():
        logger.info("Attempting advanced Windows recovery methods...")
        
        # Method 1: Use robust Excel session to create clean copy
        try:
            from .excel_session import repair_excel_workbook
            if repair_excel_workbook(file_path, fixed_path):
                logger.info(f"Successfully created repaired copy: {fixed_path}")
                return True, fixed_path
        except ImportError:
            logger.warning("Excel session module not available")
        
        # Method 2: Use pandas to extract data if file is completely locked
        try:
            import pandas as pd
            logger.info("Attempting pandas-based extraction...")
            
            # Read all sheets with pandas
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # Create a new Excel writer
            writer = pd.ExcelWriter(fixed_path, engine='openpyxl')
            
            # Copy each sheet
            sheets_copied = 0
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    sheets_copied += 1
                except Exception as sheet_e:
                    logger.warning(f"Error extracting sheet {sheet_name}: {sheet_e}")
            
            # Save the writer
            writer.close()
            
            if os.path.exists(fixed_path) and os.path.getsize(fixed_path) > 0 and sheets_copied > 0:
                logger.info(f"Successfully extracted {sheets_copied} sheets to: {fixed_path}")
                return True, fixed_path
                
        except Exception as e:
            logger.warning(f"Pandas extraction failed: {e}")
        
        # Method 3: Last resort - create a new file with basic structure
        try:
            logger.info("Attempting to create new file with basic structure...")
            
            import openpyxl
            wb = openpyxl.Workbook()
            
            # Add a note about recovery
            sheet = wb.active
            sheet.title = "Recovery"
            sheet["A1"] = "This file was created as a recovery from a locked file:"
            sheet["A2"] = file_path
            sheet["A3"] = f"Created on: {time.ctime()}"
            sheet["A4"] = "Original file was locked and could not be processed."
            
            # Save the new file
            emergency_path = fixed_path.replace("_fixed_", "_emergency_")
            wb.save(emergency_path)
            
            logger.warning(f"Created emergency recovery file: {emergency_path}")
            return False, f"Created emergency file: {emergency_path}"
            
        except Exception as e:
            logger.error(f"Emergency file creation failed: {e}")
    
    return False, "Failed to fix file-in-use error despite multiple recovery attempts"

def repair_excel_file_access(file_path: str) -> Tuple[bool, str]:
    """
    Comprehensive repair for Excel files with access issues.
    
    This function attempts multiple recovery strategies in sequence
    to fix Excel files that have access issues.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple[bool, str]: (Success status, Result message or path)
    """
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"
    
    logger.info(f"Starting comprehensive repair for: {file_path}")
    
    # Create timestamp for recovery files
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    
    # Create backup first
    backup_path = str(Path(file_path).with_stem(f"{Path(file_path).stem}_backup_{timestamp}"))
    
    # Windows-specific path normalization
    if is_windows_platform():
        try:
            from ..utils.win_path_handler import normalize_windows_path
            file_path = normalize_windows_path(file_path)
            backup_path = normalize_windows_path(backup_path)
        except ImportError:
            logger.warning("Windows path handler not available")
    
    # First, try a standard file copy to create backup
    try:
        from .file_operations import safe_file_copy
        if safe_file_copy(file_path, backup_path):
            logger.info(f"Created backup before repair: {backup_path}")
        else:
            logger.warning("Could not create backup before repair")
    except Exception as e:
        logger.warning(f"Backup creation failed: {e}")
    
    # Stage 1: Try file-in-use error fix
    success, result_path = fix_file_in_use_error(file_path)
    if success:
        return True, result_path
    
    # Stage 2: On Windows, try advanced repair methods
    if is_windows_platform():
        logger.info("Attempting deep repair methods...")
        
        # Method 1: Reset Excel automation
        try:
            from ..utils.win_api import reset_excel_automation
            reset_excel_automation()
            logger.info("Reset Excel automation settings")
            
            # Try copy again after reset
            advanced_path = str(Path(file_path).with_stem(f"{Path(file_path).stem}_repaired_{timestamp}"))
            
            from .file_operations import safe_file_copy
            if safe_file_copy(file_path, advanced_path, retries=2):
                logger.info(f"Copy successful after automation reset: {advanced_path}")
                return True, advanced_path
        except ImportError:
            logger.warning("Windows API module not available")
        
        # Method 2: Try specialized session-based repair
        try:
            from .excel_session import repair_excel_workbook
            
            advanced_path = str(Path(file_path).with_stem(f"{Path(file_path).stem}_deeprepair_{timestamp}"))
            
            if repair_excel_workbook(file_path, advanced_path):
                logger.info(f"Deep repair successful: {advanced_path}")
                return True, advanced_path
        except ImportError:
            logger.warning("Excel session module not available")
    
    # If we have a backup but all repairs failed, return the backup
    if os.path.exists(backup_path) and os.path.getsize(backup_path) > 0:
        logger.warning("All repair methods failed, using backup")
        return True, backup_path
    
    # Complete failure
    return False, "All repair methods failed"

def process_with_recovery(func: Callable, *args, **kwargs) -> Tuple[bool, Any]:
    """
    Execute a function with automatic recovery from Excel file errors.
    
    Args:
        func: Function to execute
        *args: Positional arguments for the function
        **kwargs: Keyword arguments for the function
        
    Returns:
        Tuple[bool, Any]: (Success status, Return value or error message)
    """
    try:
        # First try executing the function normally
        result = func(*args, **kwargs)
        return True, result
    except Exception as original_error:
        error_msg = str(original_error)
        logger.warning(f"Error during processing: {error_msg}")
        
        # Check if this is a file access error
        if any(error_pattern in error_msg.lower() for error_pattern in [
            "process cannot access the file", 
            "being used by another process",
            "permission denied",
            "com copy failed",
            "com validation failed",
            "excel process"
        ]):
            # File access error detected, try to recover
            logger.info("Detected file access error, attempting recovery...")
            
            # Check if first argument is a file path
            file_path = None
            if args and isinstance(args[0], str) and (
                args[0].endswith('.xlsx') or args[0].endswith('.xls')
            ):
                file_path = args[0]
            
            if file_path and os.path.exists(file_path):
                # Try to fix the file access issue
                success, fixed_path = fix_file_in_use_error(file_path)
                
                if success:
                    # Replace the file path in args
                    new_args = list(args)
                    new_args[0] = fixed_path
                    
                    try:
                        # Try executing the function with the fixed file
                        logger.info(f"Retrying with fixed file: {fixed_path}")
                        result = func(*new_args, **kwargs)
                        return True, result
                    except Exception as recovery_error:
                        logger.error(f"Recovery attempt failed: {recovery_error}")
                        return False, f"Recovery failed: {recovery_error}"
                else:
                    return False, f"Could not fix file access issues: {fixed_path}"
            else:
                return False, "Could not identify file to recover"
        
        # Not a file access error, or recovery failed
        return False, f"Error: {original_error}"