"""
Utility functions for Excel processing, particularly focusing on style handling
and error recovery mechanisms.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
import logging
from typing import Optional, Dict, Any, Tuple

logger = logging.getLogger(__name__)

def create_color_object(color_value: str) -> Color:
    """
    Create a proper Color object from an RGB string.
    
    Args:
        color_value: String RGB color (with or without alpha)
    
    Returns:
        Color: openpyxl Color object
    """
    if not color_value:
        color_value = "FFFFFF"  # Default white
        
    # Remove alpha channel if present
    if len(color_value) == 8 and color_value.startswith('FF'):
        color_value = color_value[2:]
        
    return Color(rgb=color_value)

def safe_copy_fill(src_fill) -> Optional[PatternFill]:
    """
    Safely copy a fill style with proper error handling.
    
    Args:
        src_fill: Source fill object
    
    Returns:
        PatternFill or None: New fill object or None if error occurs
    """
    if not src_fill:
        return None
        
    try:
        fill_type = src_fill.fill_type if hasattr(src_fill, 'fill_type') else 'solid'
        
        # Handle start color
        start_color = None
        if hasattr(src_fill, 'start_color') and src_fill.start_color:
            rgb_value = src_fill.start_color.rgb or "FFFFFF"
            start_color = create_color_object(rgb_value)
            
        # Handle end color
        end_color = None
        if hasattr(src_fill, 'end_color') and src_fill.end_color:
            rgb_value = src_fill.end_color.rgb or "FFFFFF"
            end_color = create_color_object(rgb_value)
            
        return PatternFill(
            fill_type=fill_type,
            start_color=start_color,
            end_color=end_color
        )
    except Exception as e:
        logger.debug(f"Error copying fill: {e}")
        return None

def safe_copy_font(src_font) -> Optional[Font]:
    """
    Safely copy a font style with proper error handling.
    
    Args:
        src_font: Source font object
    
    Returns:
        Font or None: New font object or None if error occurs
    """
    if not src_font:
        return None
        
    try:
        return Font(
            name=src_font.name,
            size=src_font.size,
            bold=src_font.bold,
            italic=src_font.italic,
            color=src_font.color,
            underline=src_font.underline if hasattr(src_font, 'underline') else None,
            strike=src_font.strike if hasattr(src_font, 'strike') else None,
            vertAlign=src_font.vertAlign if hasattr(src_font, 'vertAlign') else None
        )
    except Exception as e:
        logger.debug(f"Error copying font: {e}")
        return None

def safe_copy_border(src_border) -> Optional[Border]:
    """
    Safely copy a border style with proper error handling.
    
    Args:
        src_border: Source border object
    
    Returns:
        Border or None: New border object or None if error occurs
    """
    if not src_border:
        return None
        
    try:
        return Border(
            left=src_border.left,
            right=src_border.right,
            top=src_border.top,
            bottom=src_border.bottom,
            diagonal=src_border.diagonal if hasattr(src_border, 'diagonal') else None,
            diagonalUp=src_border.diagonalUp if hasattr(src_border, 'diagonalUp') else None,
            diagonalDown=src_border.diagonalDown if hasattr(src_border, 'diagonalDown') else None
        )
    except Exception as e:
        logger.debug(f"Error copying border: {e}")
        return None

def safe_copy_alignment(src_alignment) -> Optional[Alignment]:
    """
    Safely copy an alignment style with proper error handling.
    
    Args:
        src_alignment: Source alignment object
    
    Returns:
        Alignment or None: New alignment object or None if error occurs
    """
    if not src_alignment:
        return None
        
    try:
        return Alignment(
            horizontal=src_alignment.horizontal,
            vertical=src_alignment.vertical,
            textRotation=src_alignment.textRotation if hasattr(src_alignment, 'textRotation') else 0,
            wrapText=src_alignment.wrapText if hasattr(src_alignment, 'wrapText') else None,
            shrinkToFit=src_alignment.shrinkToFit if hasattr(src_alignment, 'shrinkToFit') else None,
            indent=src_alignment.indent if hasattr(src_alignment, 'indent') else 0,
        )
    except Exception as e:
        logger.debug(f"Error copying alignment: {e}")
        return None

def safe_copy_cell_style(src_cell, tgt_cell) -> bool:
    """
    Safely copy all styles from one cell to another with error handling.
    
    Args:
        src_cell: Source cell
        tgt_cell: Target cell
    
    Returns:
        bool: True if successful, False if any errors occurred
    """
    success = True
    
    # Copy value first
    tgt_cell.value = src_cell.value
    
    if not src_cell.has_style:
        return success
        
    # Copy each style component with individual error handling
    try:
        if font := safe_copy_font(src_cell.font):
            tgt_cell.font = font
    except Exception as e:
        logger.debug(f"Error applying font: {e}")
        success = False
        
    try:
        if fill := safe_copy_fill(src_cell.fill):
            tgt_cell.fill = fill
    except Exception as e:
        logger.debug(f"Error applying fill: {e}")
        success = False
        
    try:
        if border := safe_copy_border(src_cell.border):
            tgt_cell.border = border
    except Exception as e:
        logger.debug(f"Error applying border: {e}")
        success = False
        
    try:
        if alignment := safe_copy_alignment(src_cell.alignment):
            tgt_cell.alignment = alignment
    except Exception as e:
        logger.debug(f"Error applying alignment: {e}")
        success = False
        
    return success

def get_openpyxl_version() -> Tuple[int, int, int]:
    """
    Get openpyxl version as a tuple of integers.
    
    Returns:
        Tuple[int, int, int]: Version as (major, minor, patch)
    """
    version_str = openpyxl.__version__
    try:
        return tuple(map(int, version_str.split('.')))
    except:
        return (0, 0, 0)  # Default if parsing fails
